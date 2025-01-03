import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd
import numpy as np


wrkbk2 = openpyxl.load_workbook('Drug Data.xlsx')
sh2 = wrkbk2.active

wrkbk3 = openpyxl.load_workbook('Drug Library.xlsx')
sh3 = wrkbk3.active

drug_name = []
pdb = []
kinase_name = []
kinase_group = []
kinase_family = []

for i in range(3, sh2.max_row+1):
    drug_name = sh2.cell(row=i, column=1).value
    pdb = sh2.cell(row=i, column=3).value
    for x in range(3, sh3.max_row+1):
        if drug_name == sh3.cell(row=x, column=1).value:
            cell = sh3.cell(row=x, column=2)
            cell.value = pdb



wrkbk3.save('Drug Library.xlsx')