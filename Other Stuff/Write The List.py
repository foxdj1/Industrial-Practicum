import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd

# Open Excel Workbook that contains all the requested kinases and ligands
wrkbk = openpyxl.load_workbook("Python Source Doc.xlsx")
sh = wrkbk.active

wrkbk2 = openpyxl.load_workbook('Kinase Data.xlsx')
sh2 = wrkbk2.active

wrkbk3 = openpyxl.load_workbook('Drug Data.xlsx')
sh3 = wrkbk3.active

for i in range(2, sh2.max_row+1):
    pdb = sh2.cell(row=i, column=9).value
    kinase_group = sh2.cell(row=i, column=3).value
    kinase_family = sh2.cell(row=i, column=2).value
    kinase_name = sh2.cell(row=i, column=1).value
    kinase_pdb = sh2.cell(row=i, column=4).value
    for x in range(2, sh3.max_row+1):
        if pdb == sh3.cell(row=x, column=3).value:
            drug_name = sh3.cell(row=x, column=1).value
            row = sh.max_row + 1
            sh.cell(row=row, column=1).value = kinase_group
            sh.cell(row=row, column=2).value = kinase_family
            sh.cell(row=row, column=3).value = kinase_name
            sh.cell(row=row, column=4).value = drug_name
            sh.cell(row=row, column=5).value = pdb
            sh.cell(row=row, column=6).value = kinase_pdb

sh.cell(row=1, column=1).value = 'Kinase Group'
sh.cell(row=1, column=2).value = 'Kinase Family'
sh.cell(row=1, column=3).value = 'Kinase Name'
sh.cell(row=1, column=4).value = 'Drug Name'
sh.cell(row=1, column=5).value = 'Drug PDB'
sh.cell(row=1, column=6).value = 'Kinase PDB'


wrkbk.save('Python Source Doc.xlsx')
