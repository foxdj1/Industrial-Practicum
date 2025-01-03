from IPython.display import display, Markdown, display_markdown
import pandas as pd
import nglview as nv
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import matplotlib.pyplot as plt
import os


wrkbk = openpyxl.load_workbook('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\On Target Composites\\Overall Composite.xlsx')
rootdir = 'C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\On Target Output'
xl_files = []
for subdir, dirs, files in os.walk(rootdir):
    for file in files:
        if file.endswith('.xlsx'):
            xl_files.append(os.path.join(subdir, file))


print(xl_files)
for x in xl_files:
    wrkbk2 = openpyxl.load_workbook(x)
    ws1 = wrkbk2.active = wrkbk2['Average Residuals']
    x1 = x.split("\\", 7)
    x1 = x1[7]
    x1 = x1.rstrip('.xlsx')
    print(x1)
    ws2 = wrkbk.create_sheet(x1)
    mr = ws1.max_row
    mc = ws1.max_column
    wrkbk.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\On Target Composites\\Overall Composite.xlsx')
    for i in range(1, mr+1):
        for j in range(1, mc+1):
            c = ws1.cell(row=i, column=j).value
            if c== None:
                pass
            else:
                ws2.cell(row=i, column=j).value = c

    wrkbk.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\On Target Composites\\Overall Composite.xlsx')

