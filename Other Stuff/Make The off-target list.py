import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd
import numpy as np

def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)

# Open Excel Workbook that contains all the requested kinases and ligands
create_workbook('Drug-Off-Targets.xlsx')
wrkbk = openpyxl.load_workbook('Drug-Off-Targets.xlsx')
sh = wrkbk.active

wrkbk2 = openpyxl.load_workbook('Drug Data.xlsx')
sh2 = wrkbk2.active

wrkbk3 = openpyxl.load_workbook('Drug Library.xlsx')
sh3 = wrkbk3.active

wrkbk4 = openpyxl.load_workbook('Kinase Data.xlsx')
sh4 = wrkbk4.active

for i in range(2, sh2.max_row+1):
    for z in range(2, sh3.max_row+1):
        if sh2.cell(row=i, column=1).value == sh3.cell(row=z, column=1).value:
            for x in range(2, sh4.max_row+1):
                for y in range(6,30):
                    if sh3.cell(row=z, column=y).value == sh4.cell(row=x, column=1).value:
                        if sh4.cell(row=x, column=9).value == sh2.cell(row=i, column=3).value:
                            row=sh.max_row+1
                            sh.cell(row=row, column=1).value = sh4.cell(row=x, column=3).value
                            sh.cell(row=row, column=2).value = sh4.cell(row=x, column=2).value
                            sh.cell(row=row, column=3).value = sh4.cell(row=x, column=1).value
                            sh.cell(row=row, column=4).value = sh2.cell(row=i, column=1).value
                            sh.cell(row=row, column=5).value = sh2.cell(row=i, column=3).value
                            sh.cell(row=row, column=6).value = sh4.cell(row=x, column=4).value

sh.cell(row=1, column=1).value = 'Kinase Group'
sh.cell(row=1, column=2).value = 'Kinase Family'
sh.cell(row=1, column=3).value = 'Kinase Name'
sh.cell(row=1, column=4).value = 'Drug Name'
sh.cell(row=1, column=5).value = 'Drug PDB'
sh.cell(row=1, column=6).value = 'Kinase PDB'

wrkbk.save('Drug-Off-Targets.xlsx')




# for i in range(3, sh2.max_row+1):
#     drug_name = sh2.cell(row=i, column=1).value
#     pdb = sh2.cell(row=i, column=3).value
#     for x in range(2, sh3.max_row+1):
#         if drug_name == sh3.cell(row=x, column=1).value or pdb == sh3.cell(row=x, column=2):
#             for y in range(6, 30):
#                 kinase_name = sh3.cell(row=x, column=y).value
#                 for z in range(2, sh4.max_row+1):
#                     if kinase_name == sh4.cell(row=z, column=1).value: #and pdb == sh4.cell(row=z, column=11):
#                         kinase_group = sh4.cell(row=z, column=3).value
#                         kinase_family = sh4.cell(row=z, column=2).value
#                         drug_name = sh3.cell(row=x, column=1).value
#                         row = sh.max_row+1
#                         cell = sh.cell(row=row, column=1)
#                         cell.value = kinase_group
#                         cell = sh.cell(row=row, column=2)
#                         cell.value = kinase_family
#                         cell = sh.cell(row=row, column=3)
#                         cell.value = kinase_name
#                         cell = sh.cell(row=row, column=4)
#                         cell.value = drug_name
#                         cell = sh.cell(row=row, column=5)
#                         cell.value = pdb