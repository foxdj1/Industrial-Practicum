# Their imports
import logging
from pathlib import Path

import numpy as np

logging.getLogger("numexpr").setLevel(logging.ERROR)
from bravado.client import SwaggerClient
from bravado_core.exception import SwaggerMappingError
from IPython.display import display, Markdown, display_markdown
import pandas as pd
import nglview as nv
from rdkit import Chem
from rdkit.Chem import Draw
from rdkit.Chem.Draw import IPythonConsole, MolsToGridImage
import opencadd

# My imports
from tabulate import tabulate
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import matplotlib.pyplot as plt
import json
import os
from bs4 import BeautifulSoup
import requests
from PIL import Image
from urllib.request import urlopen
import shutil
import sys
from heapq import nsmallest
import win32com.client

#Key
#wrkbk = Drugs-Targets.xlsx                     soucedoc for kinases and their known targets
#wrkbk2 = Drugs-Targets.xlsx                    soucedoc for kinases and their known off-targets
#wrkbk3 = Target_'+on+'_Avoid_'+off+'.xlsx      doc generated to store kinase-drug pairs for desired target
#wrkbk4 = all files from source directories stored under this name
#wrkbk5 = Avoid_'+off+'_Target_'+on+'.xlsx      doc generated to store kinase-drug pairs for the undesired target
#wrkbk6 = Output.xlsx                           drugs and their scores for critical and apolar residuals


#Load List of drugs and their targets
wrkbk = openpyxl.load_workbook("Drug-Targets.xlsx")
sh = wrkbk.active

#load list of drugs and their off-targets
wrkbk2 = openpyxl.load_workbook("Drug-Off-Targets.xlsx")
sh2 = wrkbk2.active

#set directories for the on targets, off targets, and the files used to find the best drug
on_base_dir = 'C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\On Target Output\\'
off_base_dir = 'C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Off Target Output\\'
path = 'C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Runs\\'

#delete any previous runs sitting in the run folder
for filename in os.listdir(path):
    os.remove(path+filename)

#prompt for desired and undesired kinases
on = input('What Kinase are you targeting?')
off = input('What Kinase do you need to avoid?')

#verify input with user
check = input('Answer with yes or no.  '+ 'You are targeting Kinase ' + on + ' and are avoiding kinase ' + off + '.')

#if input was not verified the program will jump to else statement and end
if check == 'yes':
    
    #generate the desired target lists
    good_input_file_on_list = []
    good_input_file_off_list = []
    #generates list of drug-kinase pairs for the desired target from the on-target list
    for i in range(2, sh.max_row+1):
        if sh.cell(row=i, column=3).value == on:
            kinase_group = sh.cell(row=i, column=1).value
            kinase_group = str(kinase_group)
            kinase_family = sh.cell(row=i, column=2).value
            kinase_family = str(kinase_family)
            kinase_name = sh.cell(row=i, column=3).value
            kinase_name = str(kinase_name)
            ligand_name = sh.cell(row=i, column=4).value
            ligand_expo_id = sh.cell(row=i, column=5).value
            ligand_expo_id = str(ligand_expo_id)
            kinase_pdb_id = str(sh.cell(row=i, column=6).value)
            in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
            good_input_file_on_list.append(in_file)
    good_input_file_on_list = [*set(good_input_file_on_list)]
    #generates list of drug-kinase pairs for the desired target from the off-target list
    for i in range(2, sh2.max_row+1):
        if sh2.cell(row=i, column=3).value == on:
            kinase_group = sh2.cell(row=i, column=1).value
            kinase_group = str(kinase_group)
            kinase_family = sh2.cell(row=i, column=2).value
            kinase_family = str(kinase_family)
            kinase_name = sh2.cell(row=i, column=3).value
            kinase_name = str(kinase_name)
            ligand_name = sh2.cell(row=i, column=4).value
            ligand_expo_id = sh2.cell(row=i, column=5).value
            ligand_expo_id = str(ligand_expo_id)
            kinase_pdb_id = str(sh2.cell(row=i, column=6).value)
            in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
            good_input_file_off_list.append(in_file)
    good_input_file_off_list = [*set(good_input_file_off_list)]

    #if both the intended target lists are empty, end the code and throw an error
    if len(good_input_file_on_list) == 0 and len(good_input_file_off_list) == 0:
        sys.exit('Your target kinase is not in the data.')

    #create the undesired target lists
    bad_input_file_on_list = []
    bad_input_file_off_list = []
    #generates list of drug-kinase pairs for the undesired target from the on-target list
    for i in range(2, sh.max_row+1):
        if sh.cell(row=i, column=3).value == off:
            kinase_group = sh.cell(row=i, column=1).value
            kinase_group = str(kinase_group)
            kinase_family = sh.cell(row=i, column=2).value
            kinase_family = str(kinase_family)
            kinase_name = sh.cell(row=i, column=3).value
            kinase_name = str(kinase_name)
            ligand_name = sh.cell(row=i, column=4).value
            ligand_expo_id = sh.cell(row=i, column=5).value
            ligand_expo_id = str(ligand_expo_id)
            kinase_pdb_id = str(sh.cell(row=i, column=6).value)
            in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
            good_input_file_off_list.append(in_file)
    bad_input_file_on_list = [*set(bad_input_file_on_list)]
    #generates list of drug-kinase pairs for the undesired target from the off-target list
    for i in range(2, sh2.max_row+1):
        if sh2.cell(row=i, column=3).value == off:
            kinase_group = sh2.cell(row=i, column=1).value
            kinase_group = str(kinase_group)
            kinase_family = sh2.cell(row=i, column=2).value
            kinase_family = str(kinase_family)
            kinase_name = sh2.cell(row=i, column=3).value
            kinase_name = str(kinase_name)
            ligand_name = sh2.cell(row=i, column=4).value
            ligand_expo_id = sh2.cell(row=i, column=5).value
            ligand_expo_id = str(ligand_expo_id)
            kinase_pdb_id = str(sh2.cell(row=i, column=6).value)
            in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
            bad_input_file_off_list.append(in_file)
    bad_input_file_off_list = [*set(bad_input_file_off_list)]

    #if both the undesired lists are empty, end the code and throw error
    if len(bad_input_file_on_list) == 0 and len(bad_input_file_off_list) == 0:
        sys.exit('Your off-target kinase is not in the data.')

    #create a workbook for the intended target drug pairs
    wrkbk3 = Workbook()
    wrkbk3.save(path+'Target_'+on+'_Avoid_'+off+'.xlsx')

    #if the file has entries in it, open the excel file for the kinase-drug combination and write the
    #non-zero residuals into a new sheet in the output file.
    if len(good_input_file_on_list) > 0:
        for x in good_input_file_on_list:
            dir = on_base_dir+x+'\\'+x+'.xlsx'
            wrkbk4 = openpyxl.load_workbook(dir)
            ws = wrkbk4.active
            sh3 = wrkbk3.create_sheet(x)
            sh3.cell(row=1, column=1).value = 'Residue'
            sh3.cell(row=1, column=2).value = 'Value'
            sh3.cell(row=1, column=3).value = 'Interaction Type'
            for row in range(3, ws.max_row+1):
                for column in range(2, ws.max_column+1):
                    if float(ws.cell(row=row, column=column).value) != 0:
                        lin = sh3.max_row+1
                        sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                        sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                        sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
            wrkbk3.save(path+'Target_'+on+'_Avoid_'+off+'.xlsx')

    #if the file has entries in it, open the excel file for the kinase-drug combination and write the
    #non-zero residuals into a new sheet in the output file.
    if len(good_input_file_off_list) > 0:
        for x in good_input_file_off_list:
            dir = off_base_dir+x+'\\'+x+'.xlsx'
            wrkbk4 = openpyxl.load_workbook(dir)
            ws = wrkbk4.active
            sh3 = wrkbk3.create_sheet(x)
            sh3.cell(row=1, column=1).value = 'Residue'
            sh3.cell(row=1, column=2).value = 'Value'
            sh3.cell(row=1, column=3).value = 'Interaction Type'
            for row in range(3, ws.max_row+1):
                for column in range(2, ws.max_column+1):
                    if float(ws.cell(row=row, column=column).value) != 0:
                        lin = sh3.max_row+1
                        sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                        sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                        sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
            wrkbk3.save(path+'Target_'+on+'_Avoid_'+off+'.xlsx')

    #create the workbook fof the unintended target
    wrkbk5 = Workbook()
    wrkbk5.save(path+'Avoid_'+off+'_Target_'+on+'.xlsx')

    #if the file has entries in it, open the excel file for the kinase-drug combination and write the
    #non-zero residuals into a new sheet in the output file.
    if len(bad_input_file_on_list) > 0:
        for x in bad_input_file_on_list:
            dir = on_base_dir+x+'\\'+x+'.xlsx'
            wrkbk4 = openpyxl.load_workbook(dir)
            ws = wrkbk4.active
            sh3 = wrkbk5.create_sheet(x)
            sh3.cell(row=1, column=1).value = 'Residue'
            sh3.cell(row=1, column=2).value = 'Value'
            sh3.cell(row=1, column=3).value = 'Interaction Type'
            for row in range(3, ws.max_row+1):
                for column in range(2, ws.max_column+1):
                    if float(ws.cell(row=row, column=column).value) != 0:
                        lin = sh3.max_row+1
                        sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                        sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                        sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
            wrkbk5.save(path+'Avoid_'+off+'_Target_'+on+'.xlsx')

    #if the file has entries in it, open the excel file for the kinase-drug combination and write the
    #non-zero residuals into a new sheet in the output file.
    if len(bad_input_file_off_list) > 0:
        for x in bad_input_file_off_list:
            dir = off_base_dir + x + '\\' + x + '.xlsx'
            wrkbk4 = openpyxl.load_workbook(dir)
            ws = wrkbk4.active
            sh3 = wrkbk5.create_sheet(x)
            sh3.cell(row=1, column=1).value = 'Residue'
            sh3.cell(row=1, column=2).value = 'Value'
            sh3.cell(row=1, column=3).value = 'Interaction Type'
            for row in range(3, ws.max_row + 1):
                for column in range(2, ws.max_column + 1):
                    if float(ws.cell(row=row, column=column).value) != 0:
                        lin = sh3.max_row + 1
                        sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                        sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                        sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
            wrkbk5.save(path + 'Avoid_' + off + '_Target_' + on + '.xlsx')


    #make a list of all the drugs that bind to the intended target
    good_druglist = []
    for sheet in wrkbk3.sheetnames:
        if sheet == 'Sheet':
            continue
        drug = sheet.split('_', 3)
        drug = drug[3]
        drug = drug.strip('.xlsx')
        good_druglist.append(drug)

    # make a list of all the drugs that bind to the unintended target
    bad_druglist = []
    for sheet in wrkbk5.sheetnames:
        if sheet == 'Sheet':
            continue
        drug = sheet.split('_', 3)
        drug = drug[3]
        drug = drug.strip('.xlsx')
        bad_druglist.append(drug)

    #make a list of the drugs that bind to both the inteneded and unintended targets
    elim_drugs = []
    for x in good_druglist:
        for y in bad_druglist:
            if x==y:
                elim_drugs.append(x)

    #remove drugs that bind to both the intended and unintended targets from the intended sheet
    for sheet in wrkbk3.sheetnames:
        for x in elim_drugs:
            if elim_drugs[x] in sheet:
                wrkbk3.remove_sheet(sheet)

    #remove drugs that bind to both the intended and unintended targets from the unintended sheet
    for sheet in wrkbk5.sheetnames:
        for x in elim_drugs:
            if elim_drugs[x] in sheet:
                wrkbk5.remove_sheet(sheet)


    #set count variable to zero
    count=0
    #create a composite of all interactions between the kinase of interest and all drugs
    for sheet in wrkbk3.sheetnames:
        comp = wrkbk3.active = wrkbk3['Sheet']
        count=count+1
        #skip the empty sheet that the data will be written into
        if wrkbk3[sheet].cell(row=1, column=1).value == "":
            continue
        #for the first sheet that has data, copy all the residuals onto the composite
        if count == 2:
            for row in range(2, wrkbk3[sheet].max_row + 1):
                row1 = comp.max_row + 1
                comp.cell(row=row1, column=1, value=wrkbk3[sheet].cell(row=row, column=1).value)
                comp.cell(row=row1, column=2, value=wrkbk3[sheet].cell(row=row, column=2).value)
                comp.cell(row=row1, column=3, value=wrkbk3[sheet].cell(row=row, column=3).value)
                comp.cell(row=row1, column=4, value=1)
                wrkbk3.save(path + 'Target_' + on + '_Avoid_' + off + '.xlsx')
            continue
        #for each additional sheet, if the exact residual has already been recorded, add the frequency values together
        #if the residual has not been recorded, then add it to the sheet.
        for row in range(2, wrkbk3[sheet].max_row + 1):
            for r in range(2, 5):
                if wrkbk3[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk3[sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                    comp.cell(row=row, column=2, value=comp.cell(row=row, column=2).value + wrkbk3[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row, column=4, value=comp.cell(row=row, column=4).value + 1)
                    break
                else:
                    row1 = comp.max_row+1
                    comp.cell(row=row1, column=1, value=wrkbk3[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk3[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk3[sheet].cell(row=row, column=3).value)
                    comp.cell(row=row1, column=4, value=1)
                    wrkbk3.save(path + 'Target_' + on + '_Avoid_' + off + '.xlsx')
                    break
    wrkbk3.save(path + 'Target_' + on + '_Avoid_' + off + '.xlsx')


    comp = wrkbk3.active = wrkbk3['Sheet']
    #set length to -1, so count reflects number of non-composite sheets
    len1 = -1
    #count the sheets that have data in them
    for sheet in wrkbk3.sheetnames:
        if wrkbk3[sheet].cell(row=2, column=1).value != "":
            len1 = len1 + 1
    # divide the frequency value for each interaction by the number of drugs
    for row in range(2, comp.max_row + 1):
        comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
    wrkbk3.save(path + 'Target_' + on + '_Avoid_' + off + '.xlsx')


#set count variable to zero
    count=0
    #create a composite of all interactions between the kinase to be avoided and all drugs
    for sheet in wrkbk5.sheetnames:
        comp = wrkbk5.active = wrkbk5['Sheet']
        count=count+1
        #skip the empty sheet that the data will be written into
        if wrkbk5[sheet].cell(row=1, column=1).value == "":
            continue
        #for the first sheet that has data, copy all the residuals onto the composite
        if count == 2:
            for row in range(2, wrkbk5[sheet].max_row + 1):
                row1 = comp.max_row + 1
                comp.cell(row=row1, column=1, value=wrkbk5[sheet].cell(row=row, column=1).value)
                comp.cell(row=row1, column=2, value=wrkbk5[sheet].cell(row=row, column=2).value)
                comp.cell(row=row1, column=3, value=wrkbk5[sheet].cell(row=row, column=3).value)
                wrkbk5.save(path + 'Avoid_' + off + '_Target_' + on + '.xlsx')
            continue
        #for each additional sheet, if the exact residual has already been recorded, add the frequency values together
        #if the residual has not been recorded, then add it to the sheet.
        for row in range(2, wrkbk5[sheet].max_row + 1):
            for r in range(2, 5):
                if wrkbk5[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk5[sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                    comp.cell(row=row, column=2, value=comp.cell(row=row, column=2).value + wrkbk5[sheet].cell(row=row, column=2).value)
                    break
                else:
                    row1 = comp.max_row+1
                    comp.cell(row=row1, column=1, value=wrkbk5[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk5[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk5[sheet].cell(row=row, column=3).value)
                    wrkbk5.save(path + 'Avoid_' + off + '_Target_' + on + '.xlsx')
                    break
    wrkbk5.save(path + 'Avoid_' + off + '_Target_' + on + '.xlsx')


    comp = wrkbk5.active = wrkbk5['Sheet']
    #set length to -1, so count reflects number of non-composite sheets
    len2 = -1
    #count the sheets that have data in them
    for sheet in wrkbk5.sheetnames:
        if wrkbk5[sheet].cell(row=2, column=1).value != "":
            len2 = len2 + 1
    # divide the frequency value for each interaction by the number of drugs
    for row in range(2, comp.max_row + 1):
        comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
    wrkbk5.save(path + 'Avoid_' + off + '_Target_' + on + '.xlsx')



    comp = wrkbk5.active = wrkbk5['Sheet']
    comp2 = wrkbk3.active = wrkbk3['Sheet']
    wrkbk6 = Workbook()
    wrkbk6.save(path+'Output.xlsx')


    for sheet in wrkbk3.sheetnames:
        crit_res = 0
        crit_hits = 0
        if sheet == 'Sheet':
            continue
        drug = sheet.split('_', 3)
        drug = drug[3]
        drug = drug.strip('.xlsx')
        wb = wrkbk6.active
        wb1 = wrkbk6.create_sheet(drug)
        for row in range(2, comp.max_row+1):
            if comp.cell(row=row, column=3).value != 'Apolar contact':
                residue = comp.cell(row=row, column=1).value
                int_type = comp.cell(row=row, column=3).value
                crit_res = crit_res+1
                r = wb1.max_row+1
                wb1.cell(row=r, column=1, value=residue)
                wb1.cell(row=r, column=2, value=int_type)
                for row1 in range(2, wrkbk3[sheet].max_row+1):
                    if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1, column=3).value == int_type:
                        crit_hits = crit_hits + 1
                        wb1.cell(row=r, column=3, value=1)
                        wrkbk6.save(path + 'Output.xlsx')
        ap_res = 0
        ap_hits = 0
        for row in range(2, comp.max_row+1):
            if comp.cell(row=row, column=3).value == 'Apolar contact':
                residue = comp.cell(row=row, column=1).value
                int_type = comp.cell(row=row, column=3).value
                ap_res = ap_res+1
                for row1 in range(2, wrkbk3[sheet].max_row+1):
                    if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1, column=3).value == int_type:
                        ap_hits = ap_hits + 1

        r2 = wb.max_row+1
        wb.cell(row=r2, column=1, value=drug)
        wb.cell(row=r2, column=2, value=crit_hits/crit_res)
        wb.cell(row=r2, column=3, value=ap_hits/ap_res)
        wrkbk6.save(path + 'Output.xlsx')

    for row in wb.iter_rows(min_row=2, values_only=True):
        print(row)



#ends code if verification was not met
else:
    sys.exit('Please try again, either you answered no, or did not enter a valid yes input.')
