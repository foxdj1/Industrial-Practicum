# Their imports
import logging
from pathlib import Path

import numpy as np

logging.getLogger("numexpr").setLevel(logging.ERROR)
from bravado.client import SwaggerClient
from bravado_core.exception import SwaggerMappingError
from IPython.display import display, Markdown, display_markdown
import pandas as pd
import nglview as nv
from rdkit import Chem
from rdkit.Chem import Draw
from rdkit.Chem.Draw import IPythonConsole, MolsToGridImage
import opencadd

# My imports
from tabulate import tabulate
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import matplotlib.pyplot as plt
import json
import os
from bs4 import BeautifulSoup
import requests
from PIL import Image
from urllib.request import urlopen
import shutil

# Setup for remote access
pd.set_option("display.max_columns", 50)

#KLIFS_API_DEFINITIONS = "https://klifs.net/swagger/swagger.json"
#KLIFS_CLIENT = SwaggerClient.from_url(KLIFS_API_DEFINITIONS, config={"validate_responses": False})

from opencadd.databases.klifs import setup_remote
remote = setup_remote()
remote._client



# Open Excel Workbook that contains all the requested kinases and ligands
wrkbk = openpyxl.load_workbook("Python Source Doc.xlsx.xlsx")
sh = wrkbk.active


def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)


# Iterate through all requested kinase/ligand combinations
for i in range(2, sh.max_row+1):
    try:
        species = 'Human'
        # call each specified argument from the Excel sheet then turn them into strings so yhe API can process them
        kinase_group = sh.cell(row=i, column=1).value
        kinase_group = str(kinase_group)
        kinase_family = sh.cell(row=i, column=2).value
        kinase_family = str(kinase_family)
        kinase_name = sh.cell(row=i, column=3).value
        kinase_name = str(kinase_name)
        ligand_name = sh.cell(row=i, column=4).value
        ligand_expo_id = sh.cell(row=i, column=5).value
        ligand_expo_id = str(ligand_expo_id)
        kinase_pdb_id = str(sh.cell(row=i, column=6).value)

        name = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id \
                            + '.xlsx'
        path = 'C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\KLIFS Targets Output\\' +\
                kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
        try:
            os.mkdir(path)
        except:
            continue
        out = Workbook(name)
        out.save(path + '\\' +name)
        out = openpyxl.load_workbook(path + '\\' +name)
        ws = out.active
        ws.title = 'Average Residuals'
        ws2 = out.create_sheet('Kinase-Ligand Structure Data')
        ws3 = out.create_sheet('Bioactivity Data')

        kinases = remote.kinases.by_kinase_name(kinase_name)

        kinase_klifs_id = list(kinases["kinase.klifs_id"])


        ## Case Study 3

        # trim structure list down to only ones bound to ligand of interest
        structures = remote.structures.by_kinase_klifs_id(kinase_klifs_id)

        subset = structures[(structures["ligand.expo_id"] == ligand_expo_id)]
        subset.sort_values(by="structure.resolution")


        subset = subset.drop(
            columns=['kinase.names', 'kinase.family', 'kinase.group', 'ligand.name', 'ligand_allosteric.name',
                     'interaction.fingerprint', 'structure.filepath'])



        for row in dataframe_to_rows(subset, index=True, header=True):
            ws2.append(row)
        for column_cells in ws2.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws2.column_dimensions[column_cells[0].column_letter].width = length

        out.save(path + '\\' +name)

        ## Case Study 1
        structure_klifs_ids = list(subset['structure.klifs_id'])


        ## Case Study 2

        # get the average interaction fingerprint
        def average_n_interactions_per_residue(structure_klifs_ids):
            """
            Generate residue position x interaction type matrix that contains
            the average number of interactions per residue and interaction type.

            Parameters
            ----------
            structure_klifs_ids : list of int
                Structure KLIFS IDs.

            Returns
            -------
            pandas.DataFrame
                Average number of interactions per residue (rows) and interaction type (columns).
            """

            # Get IFP (is returned from KLIFS as string)
            ifps = remote.interactions.by_structure_klifs_id(structure_klifs_ids)
            # Split string into list of int (0, 1): structures x IFP bits matrix
            ifps = pd.DataFrame(ifps["interaction.fingerprint"].apply(lambda x: list(x)).to_list())
            ifps = ifps.astype("int32")
            # Sum up all interaction per bit position and normalize by number of IFPs
            ifp_relative = (ifps.sum() / len(ifps)).to_list()
            # Transform aggregated IFP into residue position x interaction type matrix
            residue_feature_matrix = pd.DataFrame(
                [ifp_relative[i : i + 7] for i in range(0, len(ifp_relative), 7)], index=range(1, 86)
            )
            # Add interaction type names
            columns = remote.interactions.interaction_types["interaction.name"].to_list()
            residue_feature_matrix.columns = columns
            return residue_feature_matrix

        ## Case Study 2
        residue_feature_matrix = average_n_interactions_per_residue(kinase_klifs_id)
        Markdown(
                f"Calculated average interactions per interaction type (feature) and residue position:\n\n"
                f"**Interaction types** (columns): {', '.join(residue_feature_matrix.columns)}\n\n"
                f"**Residue positions** (rows): {residue_feature_matrix.index}"
            )

        # NBVAL_CHECK_OUTPUT
        residue_feature_matrix

        for row in dataframe_to_rows(residue_feature_matrix, index=True, header=True):
            ws.append(row)
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length

        out.save(path + '\\' +name)

        residue_feature_matrix.plot.bar(
            figsize=(15, 5),
            stacked=True,
            title=kinase_name + " " + ligand_name + " Average Interaction Finger Print",
            xlabel="KLIFS pocket residue position",
            ylabel="Average number of interactions (per interaction type)",
            color=["grey", "green", "limegreen", "red", "blue", "orange", "cyan"],
        )
        plt.savefig('Average IFP.png', dpi = 75)
        img = openpyxl.drawing.image.Image('Average IFP.png')
        img.anchor = 'A90'
        ws.add_image(img)

        out.save(path + '\\' +name)

        for q in range(0, len(structure_klifs_ids)):
            coordinates = remote.coordinates.to_dataframe(structure_klifs_ids[q], entity='complex')
            coordinates = coordinates[['atom.id', 'atom.name', 'atom.x', 'atom.y', 'atom.z',
                                       'residue.id', 'residue.name', 'residue.klifs_id', 'residue.klifs_region_id']]
            coordinates = coordinates.astype(str)
            ws5 = out.create_sheet('Structure Coordinates ' + kinase_pdb_id)

            for row in dataframe_to_rows(coordinates, index=True, header=True):
                ws5.append(row)
            for column_cells in ws5.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws5.column_dimensions[column_cells[0].column_letter].width = length

            filepath = remote.coordinates.to_pdb(structure_klifs_ids[q], ".")
            shutil.move(str(filepath), str(path))


            try:
                remote.coordinates.to_pdb(structure_klifs_ids[q], ".", entity="ligand")
            except ValueError as e:
                print(e)

        out.save(path + '\\' + name)


        for z in range(0, len(structure_klifs_ids)):
            ifp = remote.interactions.by_structure_klifs_id(structure_klifs_ids[z])
            ifp = pd.DataFrame(ifp["interaction.fingerprint"].apply(lambda x: list(x)).to_list())
            ifp = ifp.astype("int32")
            ifp = (ifp.sum() / 1).to_list()
            residue_feature_matrix = pd.DataFrame(
                [ifp[i: i + 7] for i in range(0, len(ifp), 7)], index=range(1, 86)
            )
            # Add interaction type names
            columns = remote.interactions.interaction_types["interaction.name"].to_list()
            residue_feature_matrix.columns = columns
            ws4 = out.create_sheet('IFP' + kinase_pdb_id)

            for row in dataframe_to_rows(residue_feature_matrix, index=True, header=True):
                ws4.append(row)
            for column_cells in ws4.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws4.column_dimensions[column_cells[0].column_letter].width = length
            out.save(path + '\\' +name)

            residue_feature_matrix.plot.bar(
                figsize=(15, 5),
                stacked=True,
                title=kinase_name + " " + kinase_pdb_id + " "+ ligand_name + " Interaction Finger Print",
                xlabel="KLIFS pocket residue position",
                ylabel="Average number of interactions (per interaction type)",
                color=["grey", "green", "limegreen", "red", "blue", "orange", "cyan"],
            )
            plt.savefig('IFP.png', dpi=75)
            img = openpyxl.drawing.image.Image('IFP.png')
            img.anchor = 'A90'
            ws4.add_image(img)
            plt.close()

            out.save(path + '\\' +name)

        plt.close('all')
        del img



        ## Case Study 6

        # Profiling data
        bioactivities = remote.bioactivities.by_ligand_expo_id(ligand_expo_id)

        Markdown(
            f"Number of bioactivity values for {ligand_expo_id}: {len(bioactivities)}\n\n"
            f"Show example bioactivities:\n\n"
        )

        bioactivities.sort_values("ligand.bioactivity_standard_value").head()


        ACTIVITY_CUTOFF = 100
        bioactivities_active = bioactivities[
            bioactivities["ligand.bioactivity_standard_value"] < ACTIVITY_CUTOFF
        ]
        Markdown(("Number of measurements with high activity per kinase:"))
        n_bioactivities_per_target = (
            bioactivities_active.groupby("kinase.pref_name").size().sort_values(ascending=True)
            )
        n_bioactivities_per_target

        # Title over the bioactivity chart
        #display_markdown(f"Off-targets of {ligand_expo_id} based on profiling data:")
        bioactivities_active.sort_values(["ligand.bioactivity_standard_value"])

        for row in dataframe_to_rows(bioactivities_active, index=True, header=True):
            ws3.append(row)
        for column_cells in ws3.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws3.column_dimensions[column_cells[0].column_letter].width = length

        out.save(path + '\\' +name)
    except:
        print('There was an error with line' + str(i))
        pass