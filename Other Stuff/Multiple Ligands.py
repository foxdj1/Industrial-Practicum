# Their imports
import logging

import numpy as np

logging.getLogger("numexpr").setLevel(logging.ERROR)
from bravado.client import SwaggerClient
from IPython.display import display, Markdown, display_markdown
import pandas as pd
import nglview as nv
from rdkit import Chem
from rdkit.Chem.Draw import IPythonConsole, MolsToGridImage
import opencadd

# My imports
from tabulate import tabulate
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Setup for remote access
pd.set_option("display.max_columns", 50)

KLIFS_API_DEFINITIONS = "https://klifs.net/swagger/swagger.json"
KLIFS_CLIENT = SwaggerClient.from_url(KLIFS_API_DEFINITIONS, config={"validate_responses": False})

from opencadd.databases.klifs import setup_remote
session = setup_remote()

# Open Excel Workbook that contains all the requested kinases and ligands
wrkbk = openpyxl.load_workbook("Python Source Doc.xlsx")
sh = wrkbk.active

#Open Excel Workbook that contains all current ligands
wrkbk2 = openpyxl.load_workbook("Ligand Data.xlsx")
sh2 = wrkbk2.active

def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)


# Iterate through all requested kinase/ligand combinations
for i in range(2, sh.max_row+1):
    name = input('What would you like the Excel file output to be named?') + '.xlsx'
    create_workbook(name)
    out = openpyxl.load_workbook(name)
    ws = out.active
    ws.title = 'Residuals'
    ws2 = out.create_sheet('Kinase-Ligand Structure Data')
    ws3 = out.create_sheet('Bioactivity Data')

    species = 'Human'
    # call each specified argument from the excel sheet then turn yhem into strings so yhe API can process them
    kinase_group = sh.cell(row=i, column=1).value
    kinase_group = str(kinase_group)
    kinase_family = sh.cell(row=i, column=2).value
    kinase_family = str(kinase_family)
    kinase_name = sh.cell(row=i, column=3).value
    kinase_name = str(kinase_name)
    ligand_name = sh.cell(row=i, column=4).value
    # Iterate through all ligands to match the name requested to its Expo ID
    for i in range(2, sh2.max_row+1):
        if sh2.cell(row=i, column=2).value == ligand_name:
            ligand_expo_id = sh2.cell(row=i, column=3).value
    ligand_expo_id = str(ligand_expo_id)



    ## Tutorial Number 3

    # Pull a list of all the kinases in the family then iterate through that list to find the one that we want
    kinases = (
        KLIFS_CLIENT.Information.get_kinase_names(kinase_family=kinase_family, species=species)
        .response()
        .result
    )

    Markdown(
        f"Kinases in the {species.lower()} family {kinase_family} as a list of objects "
        f"that contain kinase-specific information:"
    )
    kinases

    # Use iterator to get first element of the list
    kinase_klifs_id = next(kinase.kinase_ID for kinase in kinases if kinase.name == kinase_name)
    Markdown(f"Kinase KLIFS ID for {kinase_name}: {kinase_klifs_id}")
    # NBVAL_CHECK_OUTPUT

    #ligands = KLIFS_CLIENT.Information.get_ligands(kinase_family=kinase_family, species=species)
    #ligand_expo_id = next(ligand.ligand_ID for ligand in ligands if ligand.name == ligand_name)

    ## Case Study 1

    # Return the number of structures for a specific kinase
    structures = session.structures.by_kinase_klifs_id(kinase_klifs_id)
    display_markdown(f"Number of structures for the kinase {kinase_name}: {len(structures)}")


    ## Case Study 1

    # Get the interaction fingerprint for the kinase-ligand structure
    structure_klifs_ids = structures["structure.klifs_id"].to_list()
    interaction_fingerprints = session.interactions.by_structure_klifs_id(structure_klifs_ids)
    display_markdown(
            f"Number of IFPs for {kinase_name}: {len(interaction_fingerprints)}\n\n"
        )

    interaction_fingerprints.head()


    ## Case Study 2

    # get the average interaction fingerprint
    def average_n_interactions_per_residue(structure_klifs_ids):
        """
        Generate residue position x interaction type matrix that contains
        the average number of interactions per residue and interaction type.

        Parameters
        ----------
        structure_klifs_ids : list of int
            Structure KLIFS IDs.

        Returns
        -------
        pandas.DataFrame
            Average number of interactions per residue (rows) and interaction type (columns).
        """

        # Get IFP (is returned from KLIFS as string)
        ifps = session.interactions.by_structure_klifs_id(structure_klifs_ids)
        # Split string into list of int (0, 1): structures x IFP bits matrix
        ifps = pd.DataFrame(ifps["interaction.fingerprint"].apply(lambda x: list(x)).to_list())
        ifps = ifps.astype("int32")
        # Sum up all interaction per bit position and normalize by number of IFPs
        ifp_relative = (ifps.sum() / len(ifps)).to_list()
        # Transform aggregated IFP into residue position x interaction type matrix
        residue_feature_matrix = pd.DataFrame(
            [ifp_relative[i : i + 7] for i in range(0, len(ifp_relative), 7)], index=range(1, 86)
        )
        # Add interaction type names
        columns = session.interactions.interaction_types["interaction.name"].to_list()
        residue_feature_matrix.columns = columns
        return residue_feature_matrix

    ## Case Study 2
    residue_feature_matrix = average_n_interactions_per_residue(structure_klifs_ids)
    Markdown(
            f"Calculated average interactions per interaction type (feature) and residue position:\n\n"
            f"**Interaction types** (columns): {', '.join(residue_feature_matrix.columns)}\n\n"
            f"**Residue positions** (rows): {residue_feature_matrix.index}"
        )

    # NBVAL_CHECK_OUTPUT
    residue_feature_matrix

    ## My Table code for average interaction finger print
    #print(tabulate(residue_feature_matrix, ['Residue', 'Apolar Contact', 'Aromatic face-to-face', 'Aromatic edge-to-face',
    #                                        'Hydrogen bond donor (protein)', 'Hydrogen bond acceptor (protein)',
    #                                        'Protein cation - ligand anion', 'Protein anion - ligand cation'], 'github'))



    for row in dataframe_to_rows(residue_feature_matrix, index=True, header=True):
        ws.append(row)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length


    ## Case Study 3

    # trim structure list down to only ones bound to ligand of interest
    subset = structures[(structures["ligand.expo_id"] == ligand_expo_id)]
    subset.sort_values(by="structure.resolution").head(10)

    #print(tabulate(subset, ['structure.klifs_id', 'structure.pdb_id',
    #                        'structure.alternate_model', 'structure.chain', 'species.klifs', 'kinase.klifs_id',
    #                        'kinase.klifs_name', 'kinase.names', 'kinase.family', 'kinase.group', 'structure.pocket',
    #                        'ligand.expo_id', 'ligand_allosetric.expo_id', 'ligand.klifs_id',
    #                        'ligand_allosteric.klifs_id',
    #                        'ligand.name', 'ligand_allosteric.name', 'structure.dfg', 'structure.ac_helix',
    #                        'structure.resolution', 'structure.qualityscore', 'structure.missing_residues',
    #                        'structure.missing_atoms', 'structure.rmsd1', 'structure.rmsd2', 'interaction.fingerprint',
    #                        'structure.front', 'structure.gate', 'structure.back', 'structure.fp_i', 'structure.fp_ii',
    #                        'structure.bp_i_a', 'structure.bp_i_b', 'structure.bp_ii_in', 'structure.bp_ii_a_in',
    #                        'structure.bp_ii_b_in', 'structure.bp_ii_out', 'structure.bp_ii_b', 'structure.bp_iii',
    #                        'structure.bp_iv', 'structure.bp_v', 'structure.grich_distance', 'structure.grich_angle',
    #                        'structure.grich_rotation', 'structure.filepath', 'structure.curation_flag']))

    #subset.replace(np.nan,'-',regex=True)
    subset = subset.drop(columns=['kinase.names', 'kinase.family', 'kinase.group', 'ligand.name', 'ligand_allosteric.name',
                                   'interaction.fingerprint', 'structure.filepath'])
    for row in dataframe_to_rows(subset, index=True, header=True):
      ws2.append(row)
    for column_cells in ws2.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws2.column_dimensions[column_cells[0].column_letter].width = length


    # return the ID of the best resolved ligand
    structure_pdb_id = (
        structures[(structures["ligand.expo_id"] == ligand_expo_id)]
        .sort_values(by="structure.resolution")
        .iloc[0]["structure.pdb_id"]
    )
    message = f"Structure PDB ID for best resolved {ligand_expo_id}-bound {kinase_name}: {structure_pdb_id}"
    display_markdown(message)


    ## Case Study 6

    # Profiling data
    bioactivities = session.bioactivities.by_ligand_expo_id(ligand_expo_id)

    Markdown(
        f"Number of bioactivity values for {ligand_expo_id}: {len(bioactivities)}\n\n"
        f"Show example bioactivities:\n\n"
    )

    bioactivities.sort_values("ligand.bioactivity_standard_value").head()


    ACTIVITY_CUTOFF = 100
    bioactivities_active = bioactivities[
        bioactivities["ligand.bioactivity_standard_value"] < ACTIVITY_CUTOFF
    ]
    Markdown(("Number of measurements with high activity per kinase:"))
    n_bioactivities_per_target = (
        bioactivities_active.groupby("kinase.pref_name").size().sort_values(ascending=True)
        )
    n_bioactivities_per_target

    # Title over the bioactivity chart
    #display_markdown(f"Off-targets of {ligand_expo_id} based on profiling data:")
    bioactivities_active[
        bioactivities_active['kinase.pref_name'] != "epidermal growth factor receptor erbB1"
    ].sort_values(["ligand.bioactivity_standard_value"])

    # My code for the bioactivities table
    #print(tabulate(bioactivities_active, ['kinase.pref_name', 'kinase.uniprot', 'kinase.chembl_id', 'ligand.chembl_id',
    #               'ligand.bioactivity_standard_type', 'ligand.bioactivity_standard_relation',
    #               'ligand.bioactivity_standard_value', 'ligand.bioactivity_standard_units',
    #               'ligand.bioactivity_pchembl_value', 'species.chembl', 'ligand.expo_id'], "github"))

    for row in dataframe_to_rows(bioactivities_active, index=True, header=True):
        ws3.append(row)
    for column_cells in ws3.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws3.column_dimensions[column_cells[0].column_letter].width = length

    out.save(filename=name)