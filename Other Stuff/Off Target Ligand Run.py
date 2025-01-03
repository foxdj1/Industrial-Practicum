# Their imports
import logging

import numpy as np

logging.getLogger("numexpr").setLevel(logging.ERROR)
from bravado.client import SwaggerClient
from IPython.display import display, Markdown, display_markdown
import pandas as pd
import nglview as nv
from rdkit import Chem
from rdkit.Chem.Draw import IPythonConsole, MolsToGridImage
import opencadd

# My imports
from tabulate import tabulate
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import matplotlib.pyplot as plt

# Setup for remote access
pd.set_option("display.max_columns", 50)

KLIFS_API_DEFINITIONS = "https://klifs.net/swagger/swagger.json"
KLIFS_CLIENT = SwaggerClient.from_url(KLIFS_API_DEFINITIONS, config={"validate_responses": False})

from opencadd.databases.klifs import setup_remote
session = setup_remote()

# Open Excel Workbook that contains all the requested kinases and ligands
wrkbk = openpyxl.load_workbook("Drug-Off-Targets.xlsx")
sh = wrkbk.active


def create_workbook(path):
    workbook = Workbook()
    workbook.save(path)


# Iterate through all requested kinase/ligand combinations
for i in range(2, sh.max_row+1):
    try:
        species = 'Human'
        # call each specified argument from the Excel sheet then turn them into strings so yhe API can process them
        kinase_group = sh.cell(row=i, column=1).value
        kinase_group = str(kinase_group)
        kinase_family = sh.cell(row=i, column=2).value
        kinase_family = str(kinase_family)
        kinase_name = sh.cell(row=i, column=3).value
        kinase_name = str(kinase_name)
        ligand_name = sh.cell(row=i, column=4).value
        ligand_expo_id = sh.cell(row=i, column=5).value
        ligand_expo_id = str(ligand_expo_id)
        kinase_pdb_id = str(sh.cell(row=i, column=6).value)

        name = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id \
               + '.xlsx'
        out = Workbook(name)
        out.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Off Target Output\\' + name)
        out = openpyxl.load_workbook('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Off Target Output\\' + name)
        ws = out.active
        ws.title = 'Average Residuals'
        ws2 = out.create_sheet('Kinase-Ligand Structure Data')
        ws3 = out.create_sheet('Bioactivity Data')

        ## Tutorial Number 3

        # Pull a list of all the kinases in the family then iterate through that list to find the one that we want
        kinases = (
            KLIFS_CLIENT.Information.get_kinase_names(kinase_family=kinase_family, species=species)
            .response()
            .result
        )

        Markdown(
            f"Kinases in the {species.lower()} family {kinase_family} as a list of objects "
            f"that contain kinase-specific information:"
        )
        kinases

        # Use iterator to get first element of the list
        kinase_klifs_id = next(kinase.kinase_ID for kinase in kinases if kinase.name == kinase_name)
        Markdown(f"Kinase KLIFS ID for {kinase_name}: {kinase_klifs_id}")
        # NBVAL_CHECK_OUTPUT


        ## Case Study 1

        # Return the number of structures for a specific kinase
        structures = session.structures.by_kinase_klifs_id(kinase_klifs_id)
        Markdown(f"Number of structures for the kinase {kinase_name}: {len(structures)}")

        ## Case Study 3

        # trim structure list down to only ones bound to ligand of interest
        subset = structures[(structures["ligand.expo_id"] == ligand_expo_id)]
        subset.sort_values(by="structure.resolution")

        # subset.replace(np.nan,'-',regex=True)
        subset = subset.drop(
            columns=['kinase.names', 'kinase.family', 'kinase.group', 'ligand.name', 'ligand_allosteric.name',
                     'interaction.fingerprint', 'structure.filepath'])

        for row in dataframe_to_rows(subset, index=True, header=True):
            ws2.append(row)
        for column_cells in ws2.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws2.column_dimensions[column_cells[0].column_letter].width = length

        out.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Off Target Output\\' + name)



        ## Case Study 1
        structure_klifs_ids = []
        # Get the interaction fingerprint for the kinase-ligand structure
        for x in range(2, ws2.max_row+1):
            if ws2.cell(row=x, column=3).value == kinase_pdb_id:
                structure_klifs_ids.append(ws2.cell(row=x, column=2).value)
                interaction_fingerprints = session.interactions.by_structure_klifs_id(structure_klifs_ids)
                Markdown(
                        f"Number of IFPs for {kinase_name}: {len(interaction_fingerprints)}\n\n"
                    )

                interaction_fingerprints.head()


        ## Case Study 2

        # get the average interaction fingerprint
        def average_n_interactions_per_residue(structure_klifs_ids):
            """
            Generate residue position x interaction type matrix that contains
            the average number of interactions per residue and interaction type.

            Parameters
            ----------
            structure_klifs_ids : list of int
                Structure KLIFS IDs.

            Returns
            -------
            pandas.DataFrame
                Average number of interactions per residue (rows) and interaction type (columns).
            """

            # Get IFP (is returned from KLIFS as string)
            ifps = session.interactions.by_structure_klifs_id(structure_klifs_ids)
            # Split string into list of int (0, 1): structures x IFP bits matrix
            ifps = pd.DataFrame(ifps["interaction.fingerprint"].apply(lambda x: list(x)).to_list())
            ifps = ifps.astype("int32")
            # Sum up all interaction per bit position and normalize by number of IFPs
            ifp_relative = (ifps.sum() / len(ifps)).to_list()
            # Transform aggregated IFP into residue position x interaction type matrix
            residue_feature_matrix = pd.DataFrame(
                [ifp_relative[i : i + 7] for i in range(0, len(ifp_relative), 7)], index=range(1, 86)
            )
            # Add interaction type names
            columns = session.interactions.interaction_types["interaction.name"].to_list()
            residue_feature_matrix.columns = columns
            return residue_feature_matrix

        ## Case Study 2
        residue_feature_matrix = average_n_interactions_per_residue(structure_klifs_ids)
        Markdown(
                f"Calculated average interactions per interaction type (feature) and residue position:\n\n"
                f"**Interaction types** (columns): {', '.join(residue_feature_matrix.columns)}\n\n"
                f"**Residue positions** (rows): {residue_feature_matrix.index}"
            )

        # NBVAL_CHECK_OUTPUT
        residue_feature_matrix

        for row in dataframe_to_rows(residue_feature_matrix, index=True, header=True):
            ws.append(row)
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length

        out.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Off Target Output\\' + name)
        residue_feature_matrix.plot.bar(
            figsize=(15, 5),
            stacked=True,
            title=kinase_name + " " + ligand_name + " Average Interaction Finger Print",
            xlabel="KLIFS pocket residue position",
            ylabel="Average number of interactions (per interaction type)",
            color=["grey", "green", "limegreen", "red", "blue", "orange", "cyan"],
        )
        plt.savefig('Average IFP.png', dpi=75)
        img = openpyxl.drawing.image.Image('Average IFP.png')
        img.anchor = 'A90'
        ws.add_image(img)

        out.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\On Target Output\\' + name)

        for z in range(0, len(structure_klifs_ids)):
            ifp = session.interactions.by_structure_klifs_id(structure_klifs_ids[z])
            ifp = pd.DataFrame(ifp["interaction.fingerprint"].apply(lambda x: list(x)).to_list())
            ifp = ifp.astype("int32")
            ifp = (ifp.sum() / 1).to_list()
            residue_feature_matrix = pd.DataFrame(
                [ifp[i: i + 7] for i in range(0, len(ifp), 7)], index=range(1, 86)
            )
            # Add interaction type names
            columns = session.interactions.interaction_types["interaction.name"].to_list()
            residue_feature_matrix.columns = columns
            ws4 = out.create_sheet(kinase_pdb_id)

            for row in dataframe_to_rows(residue_feature_matrix, index=True, header=True):
                ws4.append(row)
            for column_cells in ws4.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws4.column_dimensions[column_cells[0].column_letter].width = length
            out.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Off Target Output\\' + name)

            residue_feature_matrix.plot.bar(
                figsize=(15, 5),
                stacked=True,
                title=kinase_name + " " + kinase_pdb_id + " " + ligand_name + " Interaction Finger Print",
                xlabel="KLIFS pocket residue position",
                ylabel="Average number of interactions (per interaction type)",
                color=["grey", "green", "limegreen", "red", "blue", "orange", "cyan"],
            )
            plt.savefig('IFP.png', dpi=75)
            img = openpyxl.drawing.image.Image('IFP.png')
            img.anchor = 'A90'
            ws4.add_image(img)

            out.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\On Target Output\\' + name)

        plt.close('all')
        del img
        ## Case Study 6

        # Profiling data
        bioactivities = session.bioactivities.by_ligand_expo_id(ligand_expo_id)

        Markdown(
            f"Number of bioactivity values for {ligand_expo_id}: {len(bioactivities)}\n\n"
            f"Show example bioactivities:\n\n"
        )

        bioactivities.sort_values("ligand.bioactivity_standard_value").head()


        ACTIVITY_CUTOFF = 100
        bioactivities_active = bioactivities[
            bioactivities["ligand.bioactivity_standard_value"] < ACTIVITY_CUTOFF
        ]
        Markdown(("Number of measurements with high activity per kinase:"))
        n_bioactivities_per_target = (
            bioactivities_active.groupby("kinase.pref_name").size().sort_values(ascending=True)
            )
        n_bioactivities_per_target

        # Title over the bioactivity chart
        #display_markdown(f"Off-targets of {ligand_expo_id} based on profiling data:")
        bioactivities_active[
            bioactivities_active['kinase.pref_name'] != "epidermal growth factor receptor erbB1"
        ].sort_values(["ligand.bioactivity_standard_value"])

        for row in dataframe_to_rows(bioactivities_active, index=True, header=True):
            ws3.append(row)
        for column_cells in ws3.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws3.column_dimensions[column_cells[0].column_letter].width = length

        out.save('C:\\Users\\fox8fv\\PycharmProjects\\Opencadd_real\\Off Target Output\\' + name)
    except:
        print('There was an error with line' + str(i))
        pass

