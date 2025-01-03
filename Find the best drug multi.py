# Their imports
import logging
from pathlib import Path

import numpy as np

logging.getLogger("numexpr").setLevel(logging.ERROR)
#from bravado.client import SwaggerClient
#from bravado_core.exception import SwaggerMappingError
from IPython.display import display, Markdown, display_markdown
import pandas as pd
import nglview as nv
from rdkit import Chem
from rdkit.Chem import Draw
from rdkit.Chem.Draw import IPythonConsole, MolsToGridImage
#import opencadd

# My imports
from tabulate import tabulate
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import matplotlib.pyplot as plt
import json
import os
#from bs4 import BeautifulSoup
#import requests
#from PIL import Image
from urllib.request import urlopen
import shutil
import sys
from heapq import nsmallest
#import win32com.client

# Key
# wrkbk = Drugs-Targets.xlsx                     soucedoc for kinases and their known targets
# wrkbk2 = Drugs-Targets.xlsx                    soucedoc for kinases and their known off-targets
# wrkbk3 = Target_'+on[0]+'_Avoid_'+off+'.xlsx   doc generated to store kinase-drug pairs for desired target
# wrkbk4 = all files from source directories stored under this name
# wrkbk5 = Avoid_'+off[0]+'_Target_'+on+'.xlsx   doc generated to store kinase-drug pairs for the undesired target
# wrkbk6 = Output.xlsx                           drugs and their scores for critical and apolar residuals
# wrkbk7 = Target_'+on[1]+'_Avoid_'+off+'.xlsx   doc generated to store kinase-drug pairs for desired target
# wrkbk8 = Target_'+on[2]+'_Avoid_'+off+'.xlsx   doc generated to store kinase-drug pairs for desired target
# wrkbk9 = Avoid_'+off[1]+'_Target_'+on+'.xlsx   doc generated to store kinase-drug pairs for the undesired target
# wrkbk10 = Avoid_'+off[2]+'_Target_'+on+'.xlsx   doc generated to store kinase-drug pairs for the undesired target


# Load List of drugs and their targets
wrkbk = openpyxl.load_workbook("Drug-Targets.xlsx")
sh = wrkbk.active

# load list of drugs and their off-targets
wrkbk2 = openpyxl.load_workbook("Drug-Off-Targets.xlsx")
sh2 = wrkbk2.active

# set directories for the on targets, off targets, and the files used to find the best drug
on_base_dir = 'D:\\DJ_Files\\On Target Output\\'
off_base_dir = 'D:\\DJ_Files\\Off Target Output\\'
path = 'D:\\DJ_Files\\Runs\\'

# delete any previous runs sitting in the run folder
for filename in os.listdir(path):
    os.remove(path + filename)

# prompt for desired and undesired kinases
on = input('What Kinase are you targeting?(put commas between entries, max 3)')
off = input('What Kinase do you need to avoid?(put commas between entries, max 3)')

# verify input with user
check = input('Answer with yes or no.  ' + 'You are targeting Kinases ' + on + ' and are avoiding kinases ' + off + '.')
on = on.split(', ')
off = off.split(', ')


if len(on) > 3 or len(off) > 3:
    sys.exit('You submitted to many kinases in your request')

if len(on) == 1 and on[0] == '':
    sys.exit('You must submit a kinase to be targeted.')
if len(off) == 1 and off[0] == '':
    sys.exit('You must submit a kinase to be avoided.')

# if input was not verified the program will jump to else statement and end
if check == 'yes':

    good_input_file_on_list = []
    good_input_file_off_list = []

    # generates list of drug-kinase pairs for the desired target from the on-target list
    for x in on:
        good_input_file_on_list = []
        for i in range(2, sh.max_row + 1):
            if sh.cell(row=i, column=3).value == x:
                kinase_group = sh.cell(row=i, column=1).value
                kinase_group = str(kinase_group)
                kinase_family = sh.cell(row=i, column=2).value
                kinase_family = str(kinase_family)
                kinase_name = sh.cell(row=i, column=3).value
                kinase_name = str(kinase_name)
                ligand_name = sh.cell(row=i, column=4).value
                ligand_expo_id = sh.cell(row=i, column=5).value
                ligand_expo_id = str(ligand_expo_id)
                kinase_pdb_id = str(sh.cell(row=i, column=6).value)
                in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
                good_input_file_on_list.append(in_file)
        if on.index(x) == 0:
            good_input_file_on_list_1 = [*set(good_input_file_on_list)]
        if on.index(x) == 1:
            good_input_file_on_list_2 = [*set(good_input_file_on_list)]
        if on.index(x) == 2:
            good_input_file_on_list_3 = [*set(good_input_file_on_list)]


    # generates list of drug-kinase pairs for the desired target from the off-target list
    for x in on:
        good_input_file_off_list = []
        for i in range(2, sh2.max_row + 1):
            if sh2.cell(row=i, column=3).value == x:
                kinase_group = sh2.cell(row=i, column=1).value
                kinase_group = str(kinase_group)
                kinase_family = sh2.cell(row=i, column=2).value
                kinase_family = str(kinase_family)
                kinase_name = sh2.cell(row=i, column=3).value
                kinase_name = str(kinase_name)
                ligand_name = sh2.cell(row=i, column=4).value
                ligand_expo_id = sh2.cell(row=i, column=5).value
                ligand_expo_id = str(ligand_expo_id)
                kinase_pdb_id = str(sh2.cell(row=i, column=6).value)
                in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
                good_input_file_off_list.append(in_file)
        if on.index(x) == 0:
            good_input_file_off_list_1 = [*set(good_input_file_off_list)]
        if on.index(x) == 1:
            good_input_file_off_list_2 = [*set(good_input_file_off_list)]
        if on.index(x) == 2:
            good_input_file_off_list_3 = [*set(good_input_file_off_list)]


    # if both the intended target lists are empty, end the code and throw an error
    if 'good_input_file_on_list_1' in locals() or 'good_input_file_off_list_1' in locals():
        print(on[0] + ' was found in the data')
    else:
        input('Your target Kinase ' + on[0] + ' is not in the data, would you like to continue?')
        if input != 'yes':
            sys.exit('Your target kinase is not in the data.')

    if len(on) >= 2:
        if 'good_input_file_on_list_2' in locals() or 'good_input_file_off_list_2' in locals():
            print(on[1] + ' was found in the data')
        else:
            input('Your target Kinase ' + on[1] + ' is not in the data, would you like to continue?')
            if input != 'yes':
                sys.exit('Your target kinase is not in the data.')

    if len(on) == 3:
        if 'good_input_file_on_list_3' in locals() or 'good_input_file_off_list_3' in locals():
            print(on[2] + ' was found in the data')
        else:
            input('Your target Kinase ' + on[2] + ' is not in the data, would you like to continue?')
            if input != 'yes':
                sys.exit('Your target kinase is not in the data.')


    bad_input_file_on_list = []
    bad_input_file_off_list = []

    # generates list of drug-kinase pairs for the undesired target from the on-target list
    for x in off:
        bad_input_file_on_list = []
        for i in range(2, sh.max_row + 1):
            if sh.cell(row=i, column=3).value == x:
                kinase_group = sh.cell(row=i, column=1).value
                kinase_group = str(kinase_group)
                kinase_family = sh.cell(row=i, column=2).value
                kinase_family = str(kinase_family)
                kinase_name = sh.cell(row=i, column=3).value
                kinase_name = str(kinase_name)
                ligand_name = sh.cell(row=i, column=4).value
                ligand_expo_id = sh.cell(row=i, column=5).value
                ligand_expo_id = str(ligand_expo_id)
                kinase_pdb_id = str(sh.cell(row=i, column=6).value)
                in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
                bad_input_file_on_list.append(in_file)
        if off.index(x) == 0:
            bad_input_file_on_list_1 = [*set(bad_input_file_on_list)]
        if off.index(x) == 1:
            bad_input_file_on_list_2 = [*set(bad_input_file_on_list)]
        if off.index(x) == 2:
            bad_input_file_on_list_3 = [*set(bad_input_file_on_list)]


    # generates list of drug-kinase pairs for the undesired target from the off-target list
    for x in off:
        bad_input_file_off_list = []
        for i in range(2, sh2.max_row + 1):
            if sh2.cell(row=i, column=3).value == x:
                kinase_group = sh2.cell(row=i, column=1).value
                kinase_group = str(kinase_group)
                kinase_family = sh2.cell(row=i, column=2).value
                kinase_family = str(kinase_family)
                kinase_name = sh2.cell(row=i, column=3).value
                kinase_name = str(kinase_name)
                ligand_name = sh2.cell(row=i, column=4).value
                ligand_expo_id = sh2.cell(row=i, column=5).value
                ligand_expo_id = str(ligand_expo_id)
                kinase_pdb_id = str(sh2.cell(row=i, column=6).value)
                in_file = kinase_group + '_' + kinase_family + '_' + kinase_name + '_' + ligand_name + '_' + ligand_expo_id
                bad_input_file_off_list.append(in_file)
        if off.index(x) == 0:
            bad_input_file_off_list_1 = [*set(bad_input_file_off_list)]
        if off.index(x) == 1:
            bad_input_file_off_list_2 = [*set(bad_input_file_off_list)]
        if off.index(x) == 2:
            bad_input_file_off_list_3 = [*set(bad_input_file_off_list)]


    # if both the undesired lists are empty, end the code and throw error
    if 'bad_input_file_on_list_1' in locals() or 'bad_input_file_off_list_1' in locals():
        print(off[0] + ' was found in the data')
    else:
        input('Your avoided Kinase ' + off[0] + ' is not in the data, would you like to continue?')
        if input != 'yes':
            sys.exit('Your avoided kinase is not in the data.')

    if len(off) >= 2:
        if 'bad_input_file_on_list_2' in locals() or 'bad_input_file_off_list_2' in locals():
            print(off[1] + ' was found in the data')
        else:
            input('Your avoided Kinase ' + off[1] + ' is not in the data, would you like to continue?')
            if input != 'yes':
                sys.exit('Your avoided kinase is not in the data.')

    if len(off) == 3:
        if 'bad_input_file_on_list_3' in locals() or 'bad_input_file_off_list_3' in locals():
            print(off[2] + ' was found in the data')
        else:
            input('Your avoided Kinase ' + off[2] + ' is not in the data, would you like to continue?')
            if input != 'yes':
                sys.exit('Your avoided kinase is not in the data.')



# Key
# name1 = first on target
# name2 = second on target
# name3 = third on target
# name4 = first avoid target
# name5 = second avoid target
# name6 = third avoid target


    # create a workbook for the intended target drug pairs
    wrkbk3 = Workbook()
    name1 = 'Target_' + on[0] + '_' + 'Avoid_'
    for y in off:
        name1 = name1 + y + '_'
    name1 = name1 + '.xlsx'
    wrkbk3.save(path + name1)

    # if the file has entries in it, open the Excel file for the kinase-drug combination and write the
    # non-zero residuals into a new sheet in the output file.
    if 'good_input_file_on_list_1' in locals():
        if len(good_input_file_on_list_1) > 0:
            for x in good_input_file_on_list_1:
                dir = on_base_dir + x + '\\' + x + '.xlsx'
                wrkbk4 = openpyxl.load_workbook(dir)
                ws = wrkbk4.active
                sh3 = wrkbk3.create_sheet(x)
                sh3.cell(row=1, column=1).value = 'Residue'
                sh3.cell(row=1, column=2).value = 'Value'
                sh3.cell(row=1, column=3).value = 'Interaction Type'
                for row in range(3, ws.max_row + 1):
                    for column in range(2, ws.max_column + 1):
                        if float(ws.cell(row=row, column=column).value) != 0:
                            lin = sh3.max_row + 1
                            sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                            sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                            sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                wrkbk3.save(path + name1)

    # if the file has entries in it, open the excel file for the kinase-drug combination and write the
    # non-zero residuals into a new sheet in the output file.
    if 'good_input_file_off_list_1' in locals():
        if len(good_input_file_off_list_1) > 0:
            for x in good_input_file_off_list_1:
                dir = off_base_dir + x + '\\' + x + '.xlsx'
                wrkbk4 = openpyxl.load_workbook(dir)
                ws = wrkbk4.active
                sh3 = wrkbk3.create_sheet(x)
                sh3.cell(row=1, column=1).value = 'Residue'
                sh3.cell(row=1, column=2).value = 'Value'
                sh3.cell(row=1, column=3).value = 'Interaction Type'
                for row in range(3, ws.max_row + 1):
                    for column in range(2, ws.max_column + 1):
                        if float(ws.cell(row=row, column=column).value) != 0:
                            lin = sh3.max_row + 1
                            sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                            sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                            sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                wrkbk3.save(path + name1)



    if len(on) >=2:
        wrkbk7 = Workbook()
        name2 = 'Target_' + on[1] + '_' + 'Avoid_'
        for y in off:
            name2 = name2 + y + '_'
        name2 = name2 + '.xlsx'
        wrkbk7.save(path + name2)

        if 'good_input_file_on_list_2' in locals():
            if len(good_input_file_on_list_2) > 0:
                for x in good_input_file_on_list_2:
                    dir = on_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk7.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk7.save(path + name2)

        # if the file has entries in it, open the excel file for the kinase-drug combination and write the
        # non-zero residuals into a new sheet in the output file.
        if 'good_input_file_off_list_2' in locals():
            if len(good_input_file_off_list_2) > 0:
                for x in good_input_file_off_list_2:
                    dir = off_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk7.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk7.save(path + name2)

    if len(on) == 3:
        wrkbk8 = Workbook()
        name3 = 'Target_' + on[2] + '_' + 'Avoid_'
        for y in off:
            name3 = name3 + y + '_'
        name3 = name3 + '.xlsx'
        wrkbk8.save(path + name3)

        if 'good_input_file_on_list_3' in locals():
            if len(good_input_file_on_list_3) > 0:
                for x in good_input_file_on_list_3:
                    dir = on_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk8.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk8.save(path + name3)

        # if the file has entries in it, open the excel file for the kinase-drug combination and write the
        # non-zero residuals into a new sheet in the output file.
        if 'good_input_file_off_list_3' in locals():
            if len(good_input_file_off_list_3) > 0:
                for x in good_input_file_off_list_3:
                    dir = off_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk8.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk8.save(path + name3)





    # create the workbook fof the unintended target
    wrkbk5 = Workbook()
    name4 = 'Avoid_' + off[0] + '_' + 'Target_'
    for y in on:
        name4 = name4 + y + '_'
    name4 = name4 + '.xlsx'
    wrkbk5.save(path + name4)

    # if the file has entries in it, open the excel file for the kinase-drug combination and write the
    # non-zero residuals into a new sheet in the output file.
    if 'bad_input_file_on_list_1' in locals():
        if len(bad_input_file_on_list_1) > 0:
            for x in bad_input_file_on_list_1:
                dir = on_base_dir + x + '\\' + x + '.xlsx'
                wrkbk4 = openpyxl.load_workbook(dir)
                ws = wrkbk4.active
                sh3 = wrkbk5.create_sheet(x)
                sh3.cell(row=1, column=1).value = 'Residue'
                sh3.cell(row=1, column=2).value = 'Value'
                sh3.cell(row=1, column=3).value = 'Interaction Type'
                for row in range(3, ws.max_row + 1):
                    for column in range(2, ws.max_column + 1):
                        if float(ws.cell(row=row, column=column).value) != 0:
                            lin = sh3.max_row + 1
                            sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                            sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                            sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                wrkbk5.save(path + name4)

    # if the file has entries in it, open the excel file for the kinase-drug combination and write the
    # non-zero residuals into a new sheet in the output file.
    if 'bad_input_file_off_list_1' in locals():
        if len(bad_input_file_off_list_1) > 0:
            for x in bad_input_file_off_list_1:
                dir = off_base_dir + x + '\\' + x + '.xlsx'
                wrkbk4 = openpyxl.load_workbook(dir)
                ws = wrkbk4.active
                sh3 = wrkbk5.create_sheet(x)
                sh3.cell(row=1, column=1).value = 'Residue'
                sh3.cell(row=1, column=2).value = 'Value'
                sh3.cell(row=1, column=3).value = 'Interaction Type'
                for row in range(3, ws.max_row + 1):
                    for column in range(2, ws.max_column + 1):
                        if float(ws.cell(row=row, column=column).value) != 0:
                            lin = sh3.max_row + 1
                            sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                            sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                            sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                wrkbk5.save(path + name4)


    if len(off) >= 2:
        wrkbk9 = Workbook()
        name5 = 'Avoid_' + off[1] + '_' + 'Target_'
        for y in on:
            name5 = name5 + y + '_'
        name5 = name5 + '.xlsx'
        wrkbk9.save(path + name5)

        if 'bad_input_file_on_list_2' in locals():
            if len(bad_input_file_on_list_2) > 0:
                for x in bad_input_file_on_list_2:
                    dir = on_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk9.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk9.save(path + name5)

        # if the file has entries in it, open the excel file for the kinase-drug combination and write the
        # non-zero residuals into a new sheet in the output file.
        if 'bad_input_file_off_list_2' in locals():
            if len(bad_input_file_off_list_2) > 0:
                for x in bad_input_file_off_list_2:
                    dir = off_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk9.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk9.save(path + name5)


    if len(off) == 3:
        wrkbk10 = Workbook()
        name6 = 'Avoid_' + off[2] + '_' + 'Target_'
        for y in on:
            name6 = name6 + y + '_'
        name6 = name6 + '.xlsx'
        wrkbk10.save(path + name6)

        if 'bad_input_file_on_list_3' in locals():
            if len(bad_input_file_on_list_3) > 0:
                for x in bad_input_file_on_list_3:
                    dir = on_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk10.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk10.save(path + name6)

        # if the file has entries in it, open the excel file for the kinase-drug combination and write the
        # non-zero residuals into a new sheet in the output file.
        if 'bad_input_file_off_list_3' in locals():
            if len(bad_input_file_off_list_3) > 0:
                for x in bad_input_file_off_list_3:
                    dir = off_base_dir + x + '\\' + x + '.xlsx'
                    wrkbk4 = openpyxl.load_workbook(dir)
                    ws = wrkbk4.active
                    sh3 = wrkbk10.create_sheet(x)
                    sh3.cell(row=1, column=1).value = 'Residue'
                    sh3.cell(row=1, column=2).value = 'Value'
                    sh3.cell(row=1, column=3).value = 'Interaction Type'
                    for row in range(3, ws.max_row + 1):
                        for column in range(2, ws.max_column + 1):
                            if float(ws.cell(row=row, column=column).value) != 0:
                                lin = sh3.max_row + 1
                                sh3.cell(row=lin, column=1).value = ws.cell(row=row, column=1).value
                                sh3.cell(row=lin, column=2).value = ws.cell(row=row, column=column).value
                                sh3.cell(row=lin, column=3).value = ws.cell(row=1, column=column).value
                    wrkbk10.save(path + name6)







    # make a list of all the drugs that bind to the intended target
    good_druglist = []
    for sheet in wrkbk3.sheetnames:
        if sheet == 'Sheet':
            continue
        drug = sheet.split('_', 3)
        drug = drug[3]
        drug = drug.strip('.xlsx')
        good_druglist.append(drug)
    if len(on) >= 2:
        for sheet in wrkbk7.sheetnames:
            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            good_druglist.append(drug)
    if len(on) == 3:
        for sheet in wrkbk8.sheetnames:
            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            good_druglist.append(drug)

    # make a list of all the drugs that bind to the unintended target
    bad_druglist = []
    for sheet in wrkbk5.sheetnames:
        if sheet == 'Sheet':
            continue
        drug = sheet.split('_', 3)
        drug = drug[3]
        drug = drug.strip('.xlsx')
        bad_druglist.append(drug)
    if len(off) >= 2:
        for sheet in wrkbk9.sheetnames:
            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            bad_druglist.append(drug)
    if len(off) == 3:
        for sheet in wrkbk10.sheetnames:
            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            bad_druglist.append(drug)


    # make a list of the drugs that bind to both the inteneded and unintended targets
    elim_drugs = []
    for x in good_druglist:
        for y in bad_druglist:
            if x == y:
                elim_drugs.append(x)

    # remove drugs that bind to both the intended and unintended targets from the intended sheet
    for sheet in wrkbk3.sheetnames:
        for x in elim_drugs:
            if x in sheet:
                del wrkbk3[sheet]
    if len(on) >= 2:
        for sheet in wrkbk7.sheetnames:
            for x in elim_drugs:
                if x in sheet:
                    del wrkbk7[sheet]
    if len(on) == 3:
        for sheet in wrkbk8.sheetnames:
            for x in elim_drugs:
                if x in sheet:
                    del wrkbk8[sheet]

    # remove drugs that bind to both the intended and unintended targets from the unintended sheet
    for sheet in wrkbk5.sheetnames:
        for x in elim_drugs:
            if x in sheet:
                del wrkbk5[sheet]
    if len(off) >= 2:
        for sheet in wrkbk9.sheetnames:
            for x in elim_drugs:
                if x in sheet:
                    del wrkbk9[sheet]
    if len(off) == 3:
        for sheet in wrkbk10.sheetnames:
            for x in elim_drugs:
                if x in sheet:
                    del wrkbk10[sheet]


    # set count variable to zero
    count = 0
    # create a composite of all interactions between the kinase of interest and all drugs
    for sheet in wrkbk3.sheetnames:
        comp = wrkbk3.active = wrkbk3['Sheet']
        count = count + 1
        # skip the empty sheet that the data will be written into
        if wrkbk3[sheet].cell(row=1, column=1).value == "":
            continue
        # for the first sheet that has data, copy all the residuals onto the composite
        if count == 2:
            for row in range(2, wrkbk3[sheet].max_row + 1):
                row1 = comp.max_row + 1
                comp.cell(row=row1, column=1, value=wrkbk3[sheet].cell(row=row, column=1).value)
                comp.cell(row=row1, column=2, value=wrkbk3[sheet].cell(row=row, column=2).value)
                comp.cell(row=row1, column=3, value=wrkbk3[sheet].cell(row=row, column=3).value)
                comp.cell(row=row1, column=4, value=1)
                wrkbk3.save(path + name1)
            continue
        # for each additional sheet, if the exact residual has already been recorded, add the frequency values together
        # if the residual has not been recorded, then add it to the sheet.
        for row in range(2, wrkbk3[sheet].max_row + 1):
            for r in range(2, 5):
                if wrkbk3[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk3[
                    sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                    comp.cell(row=row, column=2,
                              value=comp.cell(row=row, column=2).value + wrkbk3[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row, column=4, value=comp.cell(row=row, column=4).value + 1)
                    break
                else:
                    row1 = comp.max_row + 1
                    comp.cell(row=row1, column=1, value=wrkbk3[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk3[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk3[sheet].cell(row=row, column=3).value)
                    comp.cell(row=row1, column=4, value=1)
                    wrkbk3.save(path + name1)
                    break
    wrkbk3.save(path + name1)

    comp = wrkbk3.active = wrkbk3['Sheet']
    # set length to -1, so count reflects number of non-composite sheets
    len1 = -1
    # count the sheets that have data in them
    for sheet in wrkbk3.sheetnames:
        if wrkbk3[sheet].cell(row=2, column=1).value != "":
            len1 = len1 + 1
    # divide the frequency value for each interaction by the number of drugs
    if len1 == 0:
        sys.exit("There are no drugs that can satisfy your request.")
    for row in range(2, comp.max_row + 1):
        comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
    wrkbk3.save(path + name1)


    if len(on) >= 2:
        # set count variable to zero
        count = 0
        # create a composite of all interactions between the kinase of interest and all drugs
        for sheet in wrkbk7.sheetnames:
            comp = wrkbk7.active = wrkbk7['Sheet']
            count = count + 1
            # skip the empty sheet that the data will be written into
            if wrkbk7[sheet].cell(row=1, column=1).value == "":
                continue
            # for the first sheet that has data, copy all the residuals onto the composite
            if count == 2:
                for row in range(2, wrkbk7[sheet].max_row + 1):
                    row1 = comp.max_row + 1
                    comp.cell(row=row1, column=1, value=wrkbk7[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk7[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk7[sheet].cell(row=row, column=3).value)
                    comp.cell(row=row1, column=4, value=1)
                    wrkbk7.save(path + name2)
                continue
            # for each additional sheet, if the exact residual has already been recorded, add the frequency values together
            # if the residual has not been recorded, then add it to the sheet.
            for row in range(2, wrkbk7[sheet].max_row + 1):
                for r in range(2, 5):
                    if wrkbk7[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk7[
                        sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                        comp.cell(row=row, column=2,
                                  value=comp.cell(row=row, column=2).value + wrkbk7[sheet].cell(row=row, column=2).value)
                        comp.cell(row=row, column=4, value=comp.cell(row=row, column=4).value + 1)
                        break
                    else:
                        row1 = comp.max_row + 1
                        comp.cell(row=row1, column=1, value=wrkbk7[sheet].cell(row=row, column=1).value)
                        comp.cell(row=row1, column=2, value=wrkbk7[sheet].cell(row=row, column=2).value)
                        comp.cell(row=row1, column=3, value=wrkbk7[sheet].cell(row=row, column=3).value)
                        comp.cell(row=row1, column=4, value=1)
                        wrkbk7.save(path + name2)
                        break
        wrkbk7.save(path + name2)

        comp = wrkbk7.active = wrkbk7['Sheet']
        # set length to -1, so count reflects number of non-composite sheets
        len1 = -1
        # count the sheets that have data in them
        for sheet in wrkbk7.sheetnames:
            if wrkbk7[sheet].cell(row=2, column=1).value != "":
                len1 = len1 + 1
        # divide the frequency value for each interaction by the number of drugs
        for row in range(2, comp.max_row + 1):
            comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
        wrkbk7.save(path + name2)



    if len(on) == 3:
        # set count variable to zero
        count = 0
        # create a composite of all interactions between the kinase of interest and all drugs
        for sheet in wrkbk8.sheetnames:
            comp = wrkbk8.active = wrkbk8['Sheet']
            count = count + 1
            # skip the empty sheet that the data will be written into
            if wrkbk8[sheet].cell(row=1, column=1).value == "":
                continue
            # for the first sheet that has data, copy all the residuals onto the composite
            if count == 2:
                for row in range(2, wrkbk8[sheet].max_row + 1):
                    row1 = comp.max_row + 1
                    comp.cell(row=row1, column=1, value=wrkbk8[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk8[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk8[sheet].cell(row=row, column=3).value)
                    comp.cell(row=row1, column=4, value=1)
                    wrkbk8.save(path + name3)
                continue
            # for each additional sheet, if the exact residual has already been recorded, add the frequency values together
            # if the residual has not been recorded, then add it to the sheet.
            for row in range(2, wrkbk8[sheet].max_row + 1):
                for r in range(2, 5):
                    if wrkbk8[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk8[
                        sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                        comp.cell(row=row, column=2,
                                  value=comp.cell(row=row, column=2).value + wrkbk8[sheet].cell(row=row,
                                                                                                column=2).value)
                        comp.cell(row=row, column=4, value=comp.cell(row=row, column=4).value + 1)
                        break
                    else:
                        row1 = comp.max_row + 1
                        comp.cell(row=row1, column=1, value=wrkbk8[sheet].cell(row=row, column=1).value)
                        comp.cell(row=row1, column=2, value=wrkbk8[sheet].cell(row=row, column=2).value)
                        comp.cell(row=row1, column=3, value=wrkbk8[sheet].cell(row=row, column=3).value)
                        comp.cell(row=row1, column=4, value=1)
                        wrkbk8.save(path + name3)
                        break
        wrkbk8.save(path + name3)

        comp = wrkbk8.active = wrkbk8['Sheet']
        # set length to -1, so count reflects number of non-composite sheets
        len1 = -1
        # count the sheets that have data in them
        for sheet in wrkbk8.sheetnames:
            if wrkbk8[sheet].cell(row=2, column=1).value != "":
                len1 = len1 + 1
        # divide the frequency value for each interaction by the number of drugs
        for row in range(2, comp.max_row + 1):
            comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
        wrkbk8.save(path + name3)





    # set count variable to zero
    count = 0
    # create a composite of all interactions between the kinase to be avoided and all drugs
    for sheet in wrkbk5.sheetnames:
        comp = wrkbk5.active = wrkbk5['Sheet']
        count = count + 1
        # skip the empty sheet that the data will be written into
        if wrkbk5[sheet].cell(row=1, column=1).value == "":
            continue
        # for the first sheet that has data, copy all the residuals onto the composite
        if count == 2:
            for row in range(2, wrkbk5[sheet].max_row + 1):
                row1 = comp.max_row + 1
                comp.cell(row=row1, column=1, value=wrkbk5[sheet].cell(row=row, column=1).value)
                comp.cell(row=row1, column=2, value=wrkbk5[sheet].cell(row=row, column=2).value)
                comp.cell(row=row1, column=3, value=wrkbk5[sheet].cell(row=row, column=3).value)
                wrkbk5.save(path + name4)
            continue
        # for each additional sheet, if the exact residual has already been recorded, add the frequency values together
        # if the residual has not been recorded, then add it to the sheet.
        for row in range(2, wrkbk5[sheet].max_row + 1):
            for r in range(2, 5):
                if wrkbk5[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk5[
                    sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                    comp.cell(row=row, column=2,
                              value=comp.cell(row=row, column=2).value + wrkbk5[sheet].cell(row=row, column=2).value)
                    break
                else:
                    row1 = comp.max_row + 1
                    comp.cell(row=row1, column=1, value=wrkbk5[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk5[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk5[sheet].cell(row=row, column=3).value)
                    wrkbk5.save(path + name4)
                    break
    wrkbk5.save(path + name4)

    comp = wrkbk5.active = wrkbk5['Sheet']
    # set length to -1, so count reflects number of non-composite sheets
    len2 = -1
    # count the sheets that have data in them
    for sheet in wrkbk5.sheetnames:
        if wrkbk5[sheet].cell(row=2, column=1).value != "":
            len2 = len2 + 1
    # divide the frequency value for each interaction by the number of drugs
    for row in range(2, comp.max_row + 1):
        comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
    wrkbk5.save(path + name4)


    if len(off) >= 2:
        # set count variable to zero
        count = 0
        # create a composite of all interactions between the kinase to be avoided and all drugs
        for sheet in wrkbk9.sheetnames:
            comp = wrkbk9.active = wrkbk9['Sheet']
            count = count + 1
            # skip the empty sheet that the data will be written into
            if wrkbk9[sheet].cell(row=1, column=1).value == "":
                continue
            # for the first sheet that has data, copy all the residuals onto the composite
            if count == 2:
                for row in range(2, wrkbk9[sheet].max_row + 1):
                    row1 = comp.max_row + 1
                    comp.cell(row=row1, column=1, value=wrkbk9[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk9[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk9[sheet].cell(row=row, column=3).value)
                    wrkbk9.save(path + name5)
                continue
            # for each additional sheet, if the exact residual has already been recorded, add the frequency values together
            # if the residual has not been recorded, then add it to the sheet.
            for row in range(2, wrkbk9[sheet].max_row + 1):
                for r in range(2, 5):
                    if wrkbk9[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk9[
                        sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                        comp.cell(row=row, column=2,
                                  value=comp.cell(row=row, column=2).value + wrkbk9[sheet].cell(row=row, column=2).value)
                        break
                    else:
                        row1 = comp.max_row + 1
                        comp.cell(row=row1, column=1, value=wrkbk9[sheet].cell(row=row, column=1).value)
                        comp.cell(row=row1, column=2, value=wrkbk9[sheet].cell(row=row, column=2).value)
                        comp.cell(row=row1, column=3, value=wrkbk9[sheet].cell(row=row, column=3).value)
                        wrkbk9.save(path + name5)
                        break
        wrkbk9.save(path + name5)

        comp = wrkbk9.active = wrkbk9['Sheet']
        # set length to -1, so count reflects number of non-composite sheets
        len2 = -1
        # count the sheets that have data in them
        for sheet in wrkbk9.sheetnames:
            if wrkbk9[sheet].cell(row=2, column=1).value != "":
                len2 = len2 + 1
        # divide the frequency value for each interaction by the number of drugs
        for row in range(2, comp.max_row + 1):
            comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
        wrkbk9.save(path + name5)

    if len(off) == 3:
        # set count variable to zero
        count = 0
        # create a composite of all interactions between the kinase to be avoided and all drugs
        for sheet in wrkbk10.sheetnames:
            comp = wrkbk10.active = wrkbk10['Sheet']
            count = count + 1
            # skip the empty sheet that the data will be written into
            if wrkbk10[sheet].cell(row=1, column=1).value == "":
                continue
            # for the first sheet that has data, copy all the residuals onto the composite
            if count == 2:
                for row in range(2, wrkbk10[sheet].max_row + 1):
                    row1 = comp.max_row + 1
                    comp.cell(row=row1, column=1, value=wrkbk10[sheet].cell(row=row, column=1).value)
                    comp.cell(row=row1, column=2, value=wrkbk10[sheet].cell(row=row, column=2).value)
                    comp.cell(row=row1, column=3, value=wrkbk10[sheet].cell(row=row, column=3).value)
                    wrkbk10.save(path + name6)
                continue
            # for each additional sheet, if the exact residual has already been recorded, add the frequency values together
            # if the residual has not been recorded, then add it to the sheet.
            for row in range(2, wrkbk10[sheet].max_row + 1):
                for r in range(2, 5):
                    if wrkbk10[sheet].cell(row=row, column=1).value == comp.cell(row=row, column=1).value and wrkbk10[
                        sheet].cell(row=row, column=3).value == comp.cell(row=row, column=3).value:
                        comp.cell(row=row, column=2,
                                  value=comp.cell(row=row, column=2).value + wrkbk10[sheet].cell(row=row,
                                                                                                column=2).value)
                        break
                    else:
                        row1 = comp.max_row + 1
                        comp.cell(row=row1, column=1, value=wrkbk10[sheet].cell(row=row, column=1).value)
                        comp.cell(row=row1, column=2, value=wrkbk10[sheet].cell(row=row, column=2).value)
                        comp.cell(row=row1, column=3, value=wrkbk10[sheet].cell(row=row, column=3).value)
                        wrkbk10.save(path + name6)
                        break
        wrkbk10.save(path + name6)

        comp = wrkbk10.active = wrkbk10['Sheet']
        # set length to -1, so count reflects number of non-composite sheets
        len2 = -1
        # count the sheets that have data in them
        for sheet in wrkbk10.sheetnames:
            if wrkbk10[sheet].cell(row=2, column=1).value != "":
                len2 = len2 + 1
        # divide the frequency value for each interaction by the number of drugs
        for row in range(2, comp.max_row + 1):
            comp.cell(row=row, column=2).value = comp.cell(row=row, column=2).value / len1
        wrkbk10.save(path + name6)








    wrkbk6 = Workbook()
    wrkbk6.save(path + 'Output.xlsx')

    if len(on) == 1:
        if len(off) == 1:
            comp = wrkbk5.active = wrkbk5['Sheet']
        if len(off) == 2:
            comp = wrkbk5.active = wrkbk5['Sheet']
            comp2 = wrkbk9.active = wrkbk9['Sheet']
        if len(off) == 3:
            comp = wrkbk5.active = wrkbk5['Sheet']
            comp2 = wrkbk9.active = wrkbk9['Sheet']
            comp3 = wrkbk10.active = wrkbk10['Sheet']
        for sheet in wrkbk3.sheetnames:
            crit_res = 0
            crit_hits = 0
            crit_res1 = 0
            crit_hits1 = 0
            crit_res2 = 0
            crit_hits2 = 0

            ap_res = 0
            ap_hits = 0
            ap_res1 = 0
            ap_hits1 = 0
            ap_res2 = 0
            ap_hits2 = 0
            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            wb = wrkbk6.active
            if len(off) >= 1:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[0])
                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        crit_res = crit_res + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits = crit_hits + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        ap_res = ap_res + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits = ap_hits + 1

            if len(off) >= 2:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[1])
                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        crit_res1 = crit_res1 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits1 = crit_hits1 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        ap_res1 = ap_res1 + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits1 = ap_hits1 + 1

            if len(off) == 3:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[2])
                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        crit_res2 = crit_res2 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits2 = crit_hits2 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        ap_res2 = ap_res2 + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits2 = ap_hits2 + 1

            wrkbk6.save(path + 'Output.xlsx')
            r2 = wb.max_row + 1
            wb.cell(row=1, column=1, value='Drug')

            wb.cell(row=1, column=2, value=on[0])
            wb.cell(row=r2, column=2, value=on[0])


            wb.cell(row=r2, column=1, value=drug)
            if len(off) == 1:
                wb.cell(row=1, column=3, value=off[0] +' Polar')
                wb.cell(row=1, column=4, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=3, value=crit_hits / crit_res)
                wb.cell(row=r2, column=4, value=ap_hits / ap_res)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 2:
                wb.cell(row=1, column=3, value=off[0] +' Polar')
                wb.cell(row=1, column=4, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=3, value=crit_hits / crit_res)
                wb.cell(row=r2, column=4, value=ap_hits / ap_res)
                wb.cell(row=1, column=5, value=off[1] +' Polar')
                wb.cell(row=1, column=6, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=5, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=6, value=ap_hits1 / ap_res1)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 3:
                wb.cell(row=1, column=3, value=off[0] + ' Polar')
                wb.cell(row=1, column=4, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=3, value=crit_hits / crit_res)
                wb.cell(row=r2, column=4, value=ap_hits / ap_res)
                wb.cell(row=1, column=5, value=off[1] + ' Polar')
                wb.cell(row=1, column=6, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=5, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=6, value=ap_hits1 / ap_res1)
                wb.cell(row=1, column=7, value=off[2] + ' Polar')
                wb.cell(row=1, column=8, value=off[2] + ' Apolar')
                wb.cell(row=r2, column=7, value=crit_hits2 / crit_res2)
                wb.cell(row=r2, column=8, value=ap_hits2 / ap_res2)
                wrkbk6.save(path + 'Output.xlsx')




    if len(on) == 2:
        comp4 = wrkbk3.active = wrkbk3['Sheet']
        comp5 = wrkbk7.active = wrkbk7['Sheet']
        if len(off) == 1:
            comp = wrkbk5.active = wrkbk5['Sheet']
        if len(off) == 2:
            comp = wrkbk5.active = wrkbk5['Sheet']
            comp2 = wrkbk9.active = wrkbk9['Sheet']
        if len(off) == 3:
            comp = wrkbk5.active = wrkbk5['Sheet']
            comp2 = wrkbk9.active = wrkbk9['Sheet']
            comp3 = wrkbk10.active = wrkbk10['Sheet']

        for sheet in wrkbk3.sheetnames:
            bind = ''
            crit_res = 0
            crit_hits = 0
            crit_res1 = 0
            crit_hits1 = 0
            crit_res2 = 0
            crit_hits2 = 0
            crit_res3 = 0
            crit_hits3 = 0

            ap_res = 0
            ap_hits = 0
            ap_res1 = 0
            ap_hits1 = 0
            ap_res2 = 0
            ap_hits2 = 0
            ap_res3 = 0
            ap_hits3 = 0

            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            wb = wrkbk6.active
            if len(off) >= 1:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[0])
                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        crit_res = crit_res + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits = crit_hits + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        ap_res = ap_res + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits = ap_hits + 1

            if len(off) >= 2:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[1])
                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        crit_res1 = crit_res1 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits1 = crit_hits1 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        ap_res1 = ap_res1 + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits1 = ap_hits1 + 1

            if len(off) == 3:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[2])
                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        crit_res2 = crit_res2 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits2 = crit_hits2 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        ap_res2 = ap_res2 + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits2 = ap_hits2 + 1

            if drug in wrkbk7.sheetnames:
                bind = 'yes'
            else:
                wb1 = wrkbk6.create_sheet(drug + '_' + on[1])
                for row in range(2, wrkbk3[sheet].max_row + 1):
                    if wrkbk3[sheet].cell(row=row, column=3).value != 'Apolar contact':
                        residue = wrkbk3[sheet].cell(row=row, column=1).value
                        int_type = wrkbk3[sheet].cell(row=row, column=3).value
                        crit_res3 = crit_res3 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7['Sheet'].max_row + 1):
                            if wrkbk7['Sheet'].cell(row=row1, column=1).value == residue and wrkbk7['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits3 = crit_hits3 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, wrkbk3[sheet].max_row + 1):
                    if wrkbk3[sheet].cell(row=row, column=3).value == 'Apolar contact':
                        residue = wrkbk3[sheet].cell(row=row, column=1).value
                        int_type = wrkbk3[sheet].cell(row=row, column=3).value
                        ap_res3 = ap_res3 + 1
                        for row1 in range(2, wrkbk7['Sheet'].max_row + 1):
                            if wrkbk7['Sheet'].cell(row=row1, column=1).value == residue and wrkbk7['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits3 = ap_hits3 + 1



            wrkbk6.save(path + 'Output.xlsx')
            r2 = wb.max_row + 1
            wb.cell(row=1, column=1, value='Drug')

            wb.cell(row=1, column=2, value=on[0] + " Polar")
            wb.cell(row=1, column=3, value=on[0] + " Apolar")
            wb.cell(row=1, column=4, value=on[1] + " Polar")
            wb.cell(row=1, column=5, value=on[1] + " Apolar")
            if bind == 'yes':
                wb.cell(row=r2, column=4, value=on[1])
                wb.cell(row=r2, column=5, value=on[1])
            else:
                wb.cell(row=r2, column=4, value=crit_hits3 / crit_res3)
                wb.cell(row=r2, column=5, value=ap_hits3 / ap_res3)

            wb.cell(row=r2, column=1, value=drug)
            if len(off) == 1:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 2:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wb.cell(row=1, column=8, value=off[1] + ' Polar')
                wb.cell(row=1, column=9, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=9, value=ap_hits1 / ap_res1)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 3:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wb.cell(row=1, column=8, value=off[1] + ' Polar')
                wb.cell(row=1, column=9, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=9, value=ap_hits1 / ap_res1)
                wb.cell(row=1, column=10, value=off[2] + ' Polar')
                wb.cell(row=1, column=11, value=off[2] + ' Apolar')
                wb.cell(row=r2, column=10, value=crit_hits2 / crit_res2)
                wb.cell(row=r2, column=11, value=ap_hits2 / ap_res2)
                wrkbk6.save(path + 'Output.xlsx')


        for sheet in wrkbk7.sheetnames:
            bind = ''
            crit_res = 0
            crit_hits = 0
            crit_res1 = 0
            crit_hits1 = 0
            crit_res2 = 0
            crit_hits2 = 0
            crit_res3 = 0
            crit_hits3 = 0

            ap_res = 0
            ap_hits = 0
            ap_res1 = 0
            ap_hits1 = 0
            ap_res2 = 0
            ap_hits2 = 0
            ap_res3 = 0
            ap_hits3 = 0

            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            wb = wrkbk6.active
            if len(off) >= 1:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[0])
                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        crit_res = crit_res + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits = crit_hits + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        ap_res = ap_res + 1
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits = ap_hits + 1

            if len(off) >= 2:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[1])
                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        crit_res1 = crit_res1 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,
                                    column=3).value == int_type:
                                crit_hits1 = crit_hits1 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        ap_res1 = ap_res1 + 1
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,
                                    column=3).value == int_type:
                                ap_hits1 = ap_hits1 + 1

            if len(off) == 3:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[2])
                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        crit_res2 = crit_res2 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits2 = crit_hits2 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        ap_res2 = ap_res2 + 1
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits2 = ap_hits2 + 1


            if drug in wrkbk3.sheetnames:
                bind = 'yes'
            else:
                wb1 = wrkbk6.create_sheet(drug + '_' + on[1])
                for row in range(2, wrkbk7[sheet].max_row + 1):
                    if wrkbk7[sheet].cell(row=row, column=3).value != 'Apolar contact':
                        residue = wrkbk7[sheet].cell(row=row, column=1).value
                        int_type = wrkbk7[sheet].cell(row=row, column=3).value
                        crit_res3 = crit_res3 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3['Sheet'].max_row + 1):
                            if wrkbk3['Sheet'].cell(row=row1, column=1).value == residue and wrkbk3['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits3 = crit_hits3 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, wrkbk7[sheet].max_row + 1):
                    if wrkbk7[sheet].cell(row=row, column=3).value == 'Apolar contact':
                        residue = wrkbk7[sheet].cell(row=row, column=1).value
                        int_type = wrkbk7[sheet].cell(row=row, column=3).value
                        ap_res3 = ap_res3 + 1
                        for row1 in range(2, wrkbk3['Sheet'].max_row + 1):
                            if wrkbk3['Sheet'].cell(row=row1, column=1).value == residue and wrkbk3['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits3 = ap_hits3 + 1



            wrkbk6.save(path + 'Output.xlsx')
            r2 = wb.max_row + 1
            wb.cell(row=1, column=1, value='Drug')

            wb.cell(row=1, column=2, value=on[0] + " Polar")
            wb.cell(row=1, column=3, value=on[0] + " Apolar")
            wb.cell(row=1, column=4, value=on[1] + " Polar")
            wb.cell(row=1, column=5, value=on[1] + " Apolar")

            if bind == 'yes':
                wb.cell(row=r2, column=3, value=on[0])
                wb.cell(row=r2, column=2, value=on[0])
            else:
                wb.cell(row=r2, column=2, value=crit_hits3/crit_res3)
                wb.cell(row=r2, column=3, value=ap_hits3/ap_res3)

            wb.cell(row=r2, column=1, value=drug)
            if len(off) == 1:
                wb.cell(row=1, column=6, value=off[0] +' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 2:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wb.cell(row=1, column=8, value=off[1] +' Polar')
                wb.cell(row=1, column=9, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=9, value=ap_hits1 / ap_res1)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 3:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wb.cell(row=1, column=8, value=off[1] + ' Polar')
                wb.cell(row=1, column=9, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=9, value=ap_hits1 / ap_res1)
                wb.cell(row=1, column=10, value=off[2] + ' Polar')
                wb.cell(row=1, column=11, value=off[2] + ' Apolar')
                wb.cell(row=r2, column=10, value=crit_hits2 / crit_res2)
                wb.cell(row=r2, column=11, value=ap_hits2 / ap_res2)
                wrkbk6.save(path + 'Output.xlsx')




    if len(on) == 3:
        comp4 = wrkbk3.active = wrkbk3['Sheet']
        comp5 = wrkbk7.active = wrkbk7['Sheet']
        comp6 = wrkbk8.active = wrkbk8['Sheet']
        if len(off) == 1:
            comp = wrkbk5.active = wrkbk5['Sheet']
        if len(off) == 2:
            comp = wrkbk5.active = wrkbk5['Sheet']
            comp2 = wrkbk9.active = wrkbk9['Sheet']
        if len(off) == 3:
            comp = wrkbk5.active = wrkbk5['Sheet']
            comp2 = wrkbk9.active = wrkbk9['Sheet']
            comp3 = wrkbk10.active = wrkbk10['Sheet']

        for sheet in wrkbk3.sheetnames:
            bind = ''
            crit_res = 0
            crit_hits = 0
            crit_res1 = 0
            crit_hits1 = 0
            crit_res2 = 0
            crit_hits2 = 0
            crit_res3 = 0
            crit_hits3 = 0
            crit_res4 = 0
            crit_hits4 = 0

            ap_res = 0
            ap_hits = 0
            ap_res1 = 0
            ap_hits1 = 0
            ap_res2 = 0
            ap_hits2 = 0
            ap_res3 = 0
            ap_hits3 = 0
            ap_res4 = 0
            ap_hits4 = 0

            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            wb = wrkbk6.active
            if len(off) >= 1:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[0])
                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        crit_res = crit_res + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits = crit_hits + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        ap_res = ap_res + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits = ap_hits + 1

            if len(off) >= 2:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[1])
                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        crit_res1 = crit_res1 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits1 = crit_hits1 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        ap_res1 = ap_res1 + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits1 = ap_hits1 + 1

            if len(off) == 3:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[2])
                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        crit_res2 = crit_res2 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                crit_hits2 = crit_hits2 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        ap_res2 = ap_res2 + 1
                        for row1 in range(2, wrkbk3[sheet].max_row + 1):
                            if wrkbk3[sheet].cell(row=row1, column=1).value == residue and wrkbk3[sheet].cell(row=row1,
                                                                                                              column=3).value == int_type:
                                ap_hits2 = ap_hits2 + 1

            if drug in wrkbk7.sheetnames:
                bind = 'yes'
            else:
                wb1 = wrkbk6.create_sheet(drug + '_' + on[1])
                for row in range(2, wrkbk3[sheet].max_row + 1):
                    if wrkbk3[sheet].cell(row=row, column=3).value != 'Apolar contact':
                        residue = wrkbk3[sheet].cell(row=row, column=1).value
                        int_type = wrkbk3[sheet].cell(row=row, column=3).value
                        crit_res3 = crit_res3 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7['Sheet'].max_row + 1):
                            if wrkbk7['Sheet'].cell(row=row1, column=1).value == residue and wrkbk7['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits3 = crit_hits3 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, wrkbk3[sheet].max_row + 1):
                    if wrkbk3[sheet].cell(row=row, column=3).value == 'Apolar contact':
                        residue = wrkbk3[sheet].cell(row=row, column=1).value
                        int_type = wrkbk3[sheet].cell(row=row, column=3).value
                        ap_res3 = ap_res3 + 1
                        for row1 in range(2, wrkbk7['Sheet'].max_row + 1):
                            if wrkbk7['Sheet'].cell(row=row1, column=1).value == residue and wrkbk7['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits3 = ap_hits3 + 1



            wrkbk6.save(path + 'Output.xlsx')
            r2 = wb.max_row + 1
            wb.cell(row=1, column=1, value='Drug')

            wb.cell(row=1, column=2, value=on[0] + " Polar")
            wb.cell(row=1, column=3, value=on[0] + " Apolar")
            wb.cell(row=1, column=4, value=on[1] + " Polar")
            wb.cell(row=1, column=5, value=on[1] + " Apolar")
            if bind == 'yes':
                wb.cell(row=r2, column=4, value=on[1])
                wb.cell(row=r2, column=5, value=on[1])
            else:
                wb.cell(row=r2, column=4, value=crit_hits3 / crit_res3)
                wb.cell(row=r2, column=5, value=ap_hits3 / ap_res3)

            wb.cell(row=r2, column=1, value=drug)
            if len(off) == 1:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 2:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wb.cell(row=1, column=8, value=off[1] + ' Polar')
                wb.cell(row=1, column=9, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=9, value=ap_hits1 / ap_res1)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 3:
                wb.cell(row=1, column=6, value=off[0] + ' Polar')
                wb.cell(row=1, column=7, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=6, value=crit_hits / crit_res)
                wb.cell(row=r2, column=7, value=ap_hits / ap_res)
                wb.cell(row=1, column=8, value=off[1] + ' Polar')
                wb.cell(row=1, column=9, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=9, value=ap_hits1 / ap_res1)
                wb.cell(row=1, column=10, value=off[2] + ' Polar')
                wb.cell(row=1, column=11, value=off[2] + ' Apolar')
                wb.cell(row=r2, column=10, value=crit_hits2 / crit_res2)
                wb.cell(row=r2, column=11, value=ap_hits2 / ap_res2)
                wrkbk6.save(path + 'Output.xlsx')


        for sheet in wrkbk7.sheetnames:
            bind = ''
            crit_res = 0
            crit_hits = 0
            crit_res1 = 0
            crit_hits1 = 0
            crit_res2 = 0
            crit_hits2 = 0
            crit_res3 = 0
            crit_hits3 = 0

            ap_res = 0
            ap_hits = 0
            ap_res1 = 0
            ap_hits1 = 0
            ap_res2 = 0
            ap_hits2 = 0
            ap_res3 = 0
            ap_hits3 = 0

            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            wb = wrkbk6.active
            if len(off) >= 1:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[0])
                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        crit_res = crit_res + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits = crit_hits + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        ap_res = ap_res + 1
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits = ap_hits + 1

            if len(off) >= 2:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[1])
                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        crit_res1 = crit_res1 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,
                                    column=3).value == int_type:
                                crit_hits1 = crit_hits1 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        ap_res1 = ap_res1 + 1
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,
                                    column=3).value == int_type:
                                ap_hits1 = ap_hits1 + 1

            if len(off) == 3:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[2])
                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        crit_res2 = crit_res2 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits2 = crit_hits2 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        ap_res2 = ap_res2 + 1
                        for row1 in range(2, wrkbk7[sheet].max_row + 1):
                            if wrkbk7[sheet].cell(row=row1, column=1).value == residue and wrkbk7[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits2 = ap_hits2 + 1


            if drug in wrkbk3.sheetnames or wrkbk7.sheetnames:
                bind = 'yes'
            else:
                wb1 = wrkbk6.create_sheet(drug + '_' + on[1])
                for row in range(2, wrkbk7[sheet].max_row + 1):
                    if wrkbk7[sheet].cell(row=row, column=3).value != 'Apolar contact':
                        residue = wrkbk7[sheet].cell(row=row, column=1).value
                        int_type = wrkbk7[sheet].cell(row=row, column=3).value
                        crit_res3 = crit_res3 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3['Sheet'].max_row + 1):
                            if wrkbk3['Sheet'].cell(row=row1, column=1).value == residue and wrkbk3['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits3 = crit_hits3 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, wrkbk7[sheet].max_row + 1):
                    if wrkbk7[sheet].cell(row=row, column=3).value == 'Apolar contact':
                        residue = wrkbk7[sheet].cell(row=row, column=1).value
                        int_type = wrkbk7[sheet].cell(row=row, column=3).value
                        ap_res3 = ap_res3 + 1
                        for row1 in range(2, wrkbk3['Sheet'].max_row + 1):
                            if wrkbk3['Sheet'].cell(row=row1, column=1).value == residue and wrkbk3['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits3 = ap_hits3 + 1



            wrkbk6.save(path + 'Output.xlsx')
            r2 = wb.max_row + 1
            wb.cell(row=1, column=1, value='Drug')

            wb.cell(row=1, column=2, value=on[0] + " Polar")
            wb.cell(row=1, column=3, value=on[0] + " Apolar")
            wb.cell(row=1, column=4, value=on[1] + " Polar")
            wb.cell(row=1, column=5, value=on[1] + " Apolar")
            wb.cell(row=1, column=6, value=on[2] + " Polar")
            wb.cell(row=1, column=7, value=on[2] + " Apolar")

            if bind == 'yes':
                wb.cell(row=r2, column=2, value=on[0])
                wb.cell(row=r2, column=3, value=on[0])
            else:
                wb.cell(row=r2, column=2, value=crit_hits3/crit_res3)
                wb.cell(row=r2, column=3, value=ap_hits3/ap_res3)

            wb.cell(row=r2, column=1, value=drug)
            if len(off) == 1:
                wb.cell(row=1, column=8, value=off[0] +' Polar')
                wb.cell(row=1, column=9, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits / crit_res)
                wb.cell(row=r2, column=9, value=ap_hits / ap_res)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 2:
                wb.cell(row=1, column=8, value=off[0] + ' Polar')
                wb.cell(row=1, column=9, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits / crit_res)
                wb.cell(row=r2, column=9, value=ap_hits / ap_res)
                wb.cell(row=1, column=10, value=off[1] +' Polar')
                wb.cell(row=1, column=11, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=10, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=11, value=ap_hits1 / ap_res1)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 3:
                wb.cell(row=1, column=8, value=off[0] + ' Polar')
                wb.cell(row=1, column=9, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits / crit_res)
                wb.cell(row=r2, column=9, value=ap_hits / ap_res)
                wb.cell(row=1, column=10, value=off[1] + ' Polar')
                wb.cell(row=1, column=11, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=10, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=11, value=ap_hits1 / ap_res1)
                wb.cell(row=1, column=12, value=off[2] + ' Polar')
                wb.cell(row=1, column=13, value=off[2] + ' Apolar')
                wb.cell(row=r2, column=12, value=crit_hits2 / crit_res2)
                wb.cell(row=r2, column=13, value=ap_hits2 / ap_res2)
                wrkbk6.save(path + 'Output.xlsx')

        for sheet in wrkbk8.sheetnames:
            bind = ''
            crit_res = 0
            crit_hits = 0
            crit_res1 = 0
            crit_hits1 = 0
            crit_res2 = 0
            crit_hits2 = 0
            crit_res3 = 0
            crit_hits3 = 0

            ap_res = 0
            ap_hits = 0
            ap_res1 = 0
            ap_hits1 = 0
            ap_res2 = 0
            ap_hits2 = 0
            ap_res3 = 0
            ap_hits3 = 0

            if sheet == 'Sheet':
                continue
            drug = sheet.split('_', 3)
            drug = drug[3]
            drug = drug.strip('.xlsx')
            wb = wrkbk6.active
            if len(off) >= 1:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[0])
                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        crit_res = crit_res + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk8[sheet].max_row + 1):
                            if wrkbk8[sheet].cell(row=row1, column=1).value == residue and wrkbk8[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits = crit_hits + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp.max_row + 1):
                    if comp.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp.cell(row=row, column=1).value
                        int_type = comp.cell(row=row, column=3).value
                        ap_res = ap_res + 1
                        for row1 in range(2, wrkbk8[sheet].max_row + 1):
                            if wrkbk8[sheet].cell(row=row1, column=1).value == residue and wrkbk8[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits = ap_hits + 1

            if len(off) >= 2:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[1])
                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        crit_res1 = crit_res1 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk8[sheet].max_row + 1):
                            if wrkbk8[sheet].cell(row=row1, column=1).value == residue and wrkbk8[sheet].cell(
                                    row=row1,
                                    column=3).value == int_type:
                                crit_hits1 = crit_hits1 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp2.max_row + 1):
                    if comp2.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp2.cell(row=row, column=1).value
                        int_type = comp2.cell(row=row, column=3).value
                        ap_res1 = ap_res1 + 1
                        for row1 in range(2, wrkbk8[sheet].max_row + 1):
                            if wrkbk8[sheet].cell(row=row1, column=1).value == residue and wrkbk8[sheet].cell(
                                    row=row1,
                                    column=3).value == int_type:
                                ap_hits1 = ap_hits1 + 1

            if len(off) == 3:
                wb1 = wrkbk6.create_sheet(drug + '_' + off[2])
                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value != 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        crit_res2 = crit_res2 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk8[sheet].max_row + 1):
                            if wrkbk8[sheet].cell(row=row1, column=1).value == residue and wrkbk8[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits2 = crit_hits2 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, comp3.max_row + 1):
                    if comp3.cell(row=row, column=3).value == 'Apolar contact':
                        residue = comp3.cell(row=row, column=1).value
                        int_type = comp3.cell(row=row, column=3).value
                        ap_res2 = ap_res2 + 1
                        for row1 in range(2, wrkbk8[sheet].max_row + 1):
                            if wrkbk8[sheet].cell(row=row1, column=1).value == residue and wrkbk8[sheet].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits2 = ap_hits2 + 1


            if drug in wrkbk3.sheetnames or wrkbk7.sheetnames:
                bind = 'yes'
            else:
                wb1 = wrkbk6.create_sheet(drug + '_' + on[1])
                for row in range(2, wrkbk8[sheet].max_row + 1):
                    if wrkbk8[sheet].cell(row=row, column=3).value != 'Apolar contact':
                        residue = wrkbk8[sheet].cell(row=row, column=1).value
                        int_type = wrkbk8[sheet].cell(row=row, column=3).value
                        crit_res3 = crit_res3 + 1
                        r = wb1.max_row + 1
                        wb1.cell(row=r, column=1, value=residue)
                        wb1.cell(row=r, column=2, value=int_type)
                        for row1 in range(2, wrkbk3['Sheet'].max_row + 1):
                            if wrkbk3['Sheet'].cell(row=row1, column=1).value == residue and wrkbk3['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                crit_hits3 = crit_hits3 + 1
                                wb1.cell(row=r, column=3, value=1)
                                wrkbk6.save(path + 'Output.xlsx')

                for row in range(2, wrkbk8[sheet].max_row + 1):
                    if wrkbk8[sheet].cell(row=row, column=3).value == 'Apolar contact':
                        residue = wrkbk8[sheet].cell(row=row, column=1).value
                        int_type = wrkbk8[sheet].cell(row=row, column=3).value
                        ap_res3 = ap_res3 + 1
                        for row1 in range(2, wrkbk3['Sheet'].max_row + 1):
                            if wrkbk3['Sheet'].cell(row=row1, column=1).value == residue and wrkbk3['Sheet'].cell(
                                    row=row1,column=3).value == int_type:
                                ap_hits3 = ap_hits3 + 1


            wrkbk6.save(path + 'Output.xlsx')
            r2 = wb.max_row + 1
            wb.cell(row=1, column=1, value='Drug')

            wb.cell(row=1, column=2, value=on[0] + " Polar")
            wb.cell(row=1, column=3, value=on[0] + " Apolar")
            wb.cell(row=1, column=4, value=on[1] + " Polar")
            wb.cell(row=1, column=5, value=on[1] + " Apolar")
            wb.cell(row=1, column=6, value=on[2] + " Polar")
            wb.cell(row=1, column=7, value=on[2] + " Apolar")

            if bind == 'yes':
                wb.cell(row=r2, column=2, value=on[0])
                wb.cell(row=r2, column=3, value=on[0])
            else:
                wb.cell(row=r2, column=2, value=crit_hits3 / crit_res3)
                wb.cell(row=r2, column=3, value=ap_hits3 / ap_res3)

            wb.cell(row=r2, column=1, value=drug)
            if len(off) == 1:
                wb.cell(row=1, column=8, value=off[0] + ' Polar')
                wb.cell(row=1, column=9, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits / crit_res)
                wb.cell(row=r2, column=9, value=ap_hits / ap_res)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 2:
                wb.cell(row=1, column=8, value=off[0] + ' Polar')
                wb.cell(row=1, column=9, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits / crit_res)
                wb.cell(row=r2, column=9, value=ap_hits / ap_res)
                wb.cell(row=1, column=10, value=off[1] + ' Polar')
                wb.cell(row=1, column=11, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=10, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=11, value=ap_hits1 / ap_res1)
                wrkbk6.save(path + 'Output.xlsx')
            if len(off) == 3:
                wb.cell(row=1, column=8, value=off[0] + ' Polar')
                wb.cell(row=1, column=9, value=off[0] + ' Apolar')
                wb.cell(row=r2, column=8, value=crit_hits / crit_res)
                wb.cell(row=r2, column=9, value=ap_hits / ap_res)
                wb.cell(row=1, column=10, value=off[1] + ' Polar')
                wb.cell(row=1, column=11, value=off[1] + ' Apolar')
                wb.cell(row=r2, column=10, value=crit_hits1 / crit_res1)
                wb.cell(row=r2, column=11, value=ap_hits1 / ap_res1)
                wb.cell(row=1, column=12, value=off[2] + ' Polar')
                wb.cell(row=1, column=13, value=off[2] + ' Apolar')
                wb.cell(row=r2, column=12, value=crit_hits2 / crit_res2)
                wb.cell(row=r2, column=13, value=ap_hits2 / ap_res2)
                wrkbk6.save(path + 'Output.xlsx')



    for row in wb.iter_rows(min_row=2, values_only=True):
        print(row)
    print(path + 'Output.xlsx')

    # ends code if verification was not met
else:
    sys.exit('Please try again, either you answered no, or did not enter a valid yes input.')